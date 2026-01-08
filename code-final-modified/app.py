
from flask import Flask, request, jsonify, send_file, render_template, session, redirect, url_for
from flask_cors import CORS
import oracledb
import os
from typing import List, Dict, Any
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import io
from openpyxl import load_workbook
import requests
import json
import urllib3

# Disable SSL warnings (for environments with SSL certificate issues)
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Import rate limiter for AI API calls
from rate_limiter import check_rate_limit, record_api_call, get_rate_limit_status

app = Flask(__name__)
app.secret_key = 'your-secret-key-change-this-in-production-12345'  # Change this in production

# ============================================================================
# AI MODEL CONFIGURATION
# ============================================================================

# Set your AI API key here (or use environment variable for security)
# For OpenAI: Get key from https://platform.openai.com/api-keys
# For Google Gemini: Get key from https://makersuite.google.com/app/apikey
AI_MODEL_PROVIDER = os.getenv('AI_MODEL_PROVIDER', 'gemini')  # 'openai' or 'gemini'

# Option 1: Set as environment variable (recommended for production)
# Option 2: Hardcode here for testing (NEVER commit this to git!)
AI_API_KEY = os.getenv('AI_API_KEY', '')  # Leave empty to set below

# 👇 QUICK SETUP: Paste your API key here for testing
if not AI_API_KEY:
    AI_API_KEY = 'AIzaSyAuhNWDCsHNhZDPL-hvTw4gVTRb6pcmQeo'  # ← PUT YOUR API KEY BETWEEN THE QUOTES
    # Example: AI_API_KEY = 'sk-proj-abc123...'  (for OpenAI)
    # Example: AI_API_KEY = 'AIzaSy...'  (for Google Gemini)

# OpenAI Configuration
OPENAI_API_URL = "https://api.openai.com/v1/chat/completions"
OPENAI_MODEL = "gpt-3.5-turbo"  # Using GPT-3.5 (works with free API keys)
# Note: To use GPT-4, you need a paid account with billing setup

# Google Gemini Configuration
# Using Gemini 2.5 Flash Lite (stable version) - lightweight model with separate quota
GEMINI_API_URL = "https://generativelanguage.googleapis.com/v1/models/gemini-2.5-flash-lite:generateContent"

# AI Model Availability Flag
AI_MODEL_AVAILABLE = bool(AI_API_KEY)

# Session Configuration
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(minutes=30)  # 30 minute session timeout
app.config['SESSION_REFRESH_EACH_REQUEST'] = True  # Reset timeout on each request

CORS(app)

# Setup logging first
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Multi-database support imports with graceful fallback
DATABRICKS_AVAILABLE = False
SNOWFLAKE_AVAILABLE = False

try:
    from databricks import sql as databricks_sql
    DATABRICKS_AVAILABLE = True
    logger.info("Databricks connector loaded successfully")
except ImportError:
    logger.warning("Databricks connector not available. Install with: pip install databricks-sql-connector")
    databricks_sql = None

try:
    import snowflake.connector
    SNOWFLAKE_AVAILABLE = True
    logger.info("Snowflake connector loaded successfully")
except ImportError:
    logger.warning("Snowflake connector not available. Install with: pip install snowflake-connector-python")
    snowflake = None

# ============================================================================
# CONFIGURATION - Update these values for your environment
# ============================================================================

# Oracle Client Configuration
# Path to Oracle Instant Client library (required for thick mode)
ORACLE_CLIENT_LIB_DIR = r"C:\oracle\instantclient_21_19"

# Backup Directory - Where exported Excel files will be saved
BACKUP_DIR = r"C:\Users\jadhabhi\Downloads\poc_genai"

# Cloud deployment flag - set to True when deploying to cloud platforms
# When True, uses thin mode (no Oracle Instant Client needed)
IS_CLOUD_DEPLOYMENT = os.getenv('IS_CLOUD_DEPLOYMENT', 'false').lower() == 'true'

# ============================================================================

# Initialize Oracle Client in thick mode with explicit path (only for local development)
if not IS_CLOUD_DEPLOYMENT:
    try:
        oracledb.init_oracle_client(lib_dir=ORACLE_CLIENT_LIB_DIR)
        logger.info("Oracle client initialized in thick mode")
    except Exception as e:
        logger.warning(f"Could not initialize thick mode: {e}")
        logger.warning("Falling back to thin mode - some Oracle features may be limited")
else:
    logger.info("Cloud deployment detected - using Oracle thin mode (no Instant Client required)")

# Create backup directory if it doesn't exist
os.makedirs(BACKUP_DIR, exist_ok=True)

# Display available database connectors
def log_available_connectors():
    """Log which database connectors are available"""
    logger.info("=" * 60)
    logger.info("DATABASE CONNECTOR STATUS")
    logger.info("=" * 60)
    logger.info(f"Oracle Database: ✓ Available (oracledb installed)")
    logger.info(f"Databricks: {'✓ Available' if DATABRICKS_AVAILABLE else '✗ Not Available'}")
    logger.info(f"Snowflake: {'✓ Available' if SNOWFLAKE_AVAILABLE else '✗ Not Available'}")
    logger.info("=" * 60)
    
    if not DATABRICKS_AVAILABLE or not SNOWFLAKE_AVAILABLE:
        logger.info("To install missing connectors:")
        if not DATABRICKS_AVAILABLE:
            logger.info("  Databricks: pip install databricks-sql-connector")
        if not SNOWFLAKE_AVAILABLE:
            logger.info("  Snowflake: pip install snowflake-connector-python")
        logger.info("  Or install all: pip install -r requirements.txt")
        logger.info("=" * 60)

log_available_connectors()

# Display startup banner
def display_startup_banner():
    """Display application startup banner"""
    banner = """
╔══════════════════════════════════════════════════════════════════╗
║                                                                  ║
║                    CrossDB Analyzer v2.0                        ║
║          Cross-Platform Database Analysis & Comparison          ║
║                                                                  ║
║              Developed by Amdocs Data Team                      ║
║                                                                  ║
║  Supported Databases: Oracle | Databricks | Snowflake          ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝
    """
    logger.info(banner)

display_startup_banner()

# Connection cache to prevent re-authentication (especially for Databricks Azure AD)
_connection_cache = {}

def get_db_connection():
    """
    Establish database connection using session credentials.
    Supports Oracle, Databricks, and Snowflake databases.
    
    This function retrieves database credentials from the Flask session and creates
    a connection to the appropriate database based on the db_type stored in session.
    
    Returns:
        Connection: An active database connection object (Oracle/Databricks/Snowflake)
            that can be used to create cursors and execute queries.
    
    Raises:
        Exception: If database credentials or db_type are not found in the session,
            or if the database type is not supported.
        DatabaseError: If there is an error establishing the database connection.
    
    Example:
        >>> connection = get_db_connection()
        >>> cursor = connection.cursor()
        >>> # ... perform database operations
        >>> connection.close()
    """
    try:
        # Check for cached connection first (prevents re-auth for Azure AD/SSO)
        session_id = session.get('session_id')
        if session_id and session_id in _connection_cache:
            cached_conn = _connection_cache[session_id]
            try:
                # Test if connection is still valid
                cursor = cached_conn.cursor()
                cursor.execute("SELECT 1")
                cursor.close()
                logger.info("Reusing cached database connection")
                return cached_conn
            except:
                # Connection is dead, remove from cache
                logger.warning("Cached connection is dead, creating new connection")
                del _connection_cache[session_id]
        
        # Check if session has required credentials
        if 'db_type' not in session:
            raise Exception("No database type in session. Please login first.")
        
        db_type = session['db_type']
        
        # Oracle Database Connection
        if db_type == 'oracle':
            if 'db_user' not in session or 'db_password' not in session or 'db_dsn' not in session:
                raise Exception("Missing Oracle credentials in session. Please login first.")
            
            connection = oracledb.connect(
                user=session['db_user'],
                password=session['db_password'],
                dsn=session['db_dsn']
            )
            return connection
        
        # Databricks Connection
        elif db_type == 'databricks':
            if not DATABRICKS_AVAILABLE:
                raise Exception("Databricks connector not installed. Install with: pip install databricks-sql-connector")
            
            if 'db_server_hostname' not in session or 'db_http_path' not in session:
                raise Exception("Missing Databricks server information in session. Please login first.")
            
            # Support multiple authentication methods
            if session.get('db_authenticator') == 'azuread':
                # Method 1: Azure AD / Entra ID authentication
                connection = databricks_sql.connect(
                    server_hostname=session['db_server_hostname'],
                    http_path=session['db_http_path'],
                    auth_type='azure-ad'
                )
            elif 'db_access_token' in session and session['db_access_token']:
                # Method 2: Token-based authentication
                connection = databricks_sql.connect(
                    server_hostname=session['db_server_hostname'],
                    http_path=session['db_http_path'],
                    access_token=session['db_access_token']
                )
            elif 'db_user' in session and 'db_password' in session:
                # Method 3: Username/Password authentication
                connection = databricks_sql.connect(
                    server_hostname=session['db_server_hostname'],
                    http_path=session['db_http_path'],
                    username=session['db_user'],
                    password=session['db_password']
                )
            else:
                raise Exception("Missing Databricks authentication. Provide Access Token, Username/Password, or use Azure AD.")
            
            return connection
        
        # Snowflake Connection
        elif db_type == 'snowflake':
            if not SNOWFLAKE_AVAILABLE:
                raise Exception("Snowflake connector not installed. Install with: pip install snowflake-connector-python")
            
            if 'db_user' not in session or 'db_account' not in session:
                raise Exception("Missing Snowflake credentials in session. Please login first.")
            
            # Build connection parameters
            conn_params = {
                'user': session['db_user'],
                'account': session['db_account']
            }
            
            # Add optional parameters
            if session.get('db_warehouse'):
                conn_params['warehouse'] = session['db_warehouse']
            if session.get('db_database'):
                conn_params['database'] = session['db_database']
            if session.get('db_schema'):
                conn_params['schema'] = session['db_schema']
            
            # Support multiple authentication methods
            if session.get('db_authenticator') == 'externalbrowser':
                # SSO/Browser-based authentication
                conn_params['authenticator'] = 'externalbrowser'
                conn_params['login_timeout'] = 120  # 2 minutes timeout for browser login
            elif session.get('db_password'):
                # Standard username/password authentication
                conn_params['password'] = session['db_password']
            else:
                raise Exception("Missing Snowflake authentication. Provide either password or use SSO.")
            
            logger.info(f"Connecting to Snowflake with authenticator: {conn_params.get('authenticator', 'password')}")
            connection = snowflake.connector.connect(**conn_params)
            return connection
        
        else:
            raise Exception(f"Unsupported database type: {db_type}")
            
    except Exception as error:
        logger.error(f"Database connection error: {error}")
        raise

def get_db_type():
    """Get the current database type from session"""
    return session.get('db_type', 'oracle')

def parse_table_input(table_input: str) -> tuple:
    """
    Parse table input to extract catalog/database, schema, and table name.
    Supports multiple formats based on database type:
    
    Oracle:
    - TABLE_NAME (uses default schema)
    - SCHEMA.TABLE_NAME
    
    Databricks:
    - TABLE_NAME (uses current catalog and schema)
    - SCHEMA.TABLE_NAME (uses current catalog)
    - CATALOG.SCHEMA.TABLE_NAME
    
    Snowflake:
    - TABLE_NAME (uses current database and schema)
    - SCHEMA.TABLE_NAME (uses current database)
    - DATABASE.SCHEMA.TABLE_NAME
    
    Returns:
        tuple: (catalog_or_database, schema, table_name)
               For Oracle: (None, schema, table)
               For Databricks: (catalog, schema, table)
               For Snowflake: (database, schema, table)
    """
    table_input = table_input.strip()
    
    # Count dots to determine format
    dot_count = table_input.count('.')
    
    if dot_count == 0:
        # Just table name - TABLE_NAME
        return None, None, table_input.upper()
    elif dot_count == 1:
        # Two-part name - SCHEMA.TABLE_NAME
        parts = table_input.split('.')
        return None, parts[0].strip().upper(), parts[1].strip().upper()
    elif dot_count == 2:
        # Three-part name - CATALOG/DATABASE.SCHEMA.TABLE_NAME
        parts = table_input.split('.')
        return parts[0].strip().upper(), parts[1].strip().upper(), parts[2].strip().upper()
    else:
        # More than 2 dots - invalid, treat as table name
        logger.warning(f"Invalid table name format: {table_input}. Treating as table name only.")
        return None, None, table_input.upper()

def get_table_count(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> Dict[str, Any]:
    """
    Get row count for a table.
    
    Args:
        cursor: Database cursor
        table_name: Table name
        schema: Schema/owner name (optional)
        catalog_or_db: Catalog (Databricks) or Database (Snowflake) name (optional)
    
    Returns:
        Dict with table_name, schema, catalog_or_db, row_count, and optional error
    """
    try:
        db_type = get_db_type()
        full_table_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        
        query = f"SELECT COUNT(*) FROM {full_table_name}"
        cursor.execute(query)
        count = cursor.fetchone()[0]
        
        return {
            'table_name': table_name,
            'schema': schema,
            'catalog_or_db': catalog_or_db,
            'owner': schema,  # For backward compatibility
            'row_count': count
        }
    except Exception as error:
        logger.error(f"Error getting count for {full_table_name if 'full_table_name' in locals() else table_name}: {error}")
        return {
            'table_name': table_name,
            'schema': schema,
            'catalog_or_db': catalog_or_db,
            'owner': schema,
            'row_count': None,
            'error': str(error)
        }

def build_full_table_name(table_name: str, schema: str = None, catalog_or_db: str = None, db_type: str = None) -> str:
    """
    Build fully qualified table name based on database type.
    
    Args:
        table_name: Table name
        schema: Schema/owner name
        catalog_or_db: Catalog (Databricks) or Database (Snowflake)
        db_type: Database type (oracle, databricks, snowflake)
    
    Returns:
        Fully qualified table name
    """
    if db_type is None:
        db_type = get_db_type()
    
    if db_type == 'oracle':
        # Oracle: SCHEMA.TABLE or just TABLE
        if schema:
            return f"{schema}.{table_name}"
        else:
            return table_name
    
    elif db_type == 'databricks':
        # Databricks: CATALOG.SCHEMA.TABLE or SCHEMA.TABLE or just TABLE
        if catalog_or_db and schema:
            return f"{catalog_or_db}.{schema}.{table_name}"
        elif schema:
            return f"{schema}.{table_name}"
        else:
            return table_name
    
    elif db_type == 'snowflake':
        # Snowflake: DATABASE.SCHEMA.TABLE or SCHEMA.TABLE or just TABLE
        if catalog_or_db and schema:
            return f"{catalog_or_db}.{schema}.{table_name}"
        elif schema:
            # Use session database if available
            if 'db_database' in session and session['db_database']:
                return f"{session['db_database']}.{schema}.{table_name}"
            else:
                return f"{schema}.{table_name}"
        else:
            return table_name
    
    else:
        # Fallback
        if schema:
            return f"{schema}.{table_name}"
        return table_name

def get_table_structure(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> List[Dict[str, Any]]:
    """Get table structure (columns) - Multi-database support"""
    try:
        db_type = get_db_type()
        structure = []
        
        if db_type == 'oracle':
            # Oracle query
            query = """
                SELECT 
                    COLUMN_NAME,
                    DATA_TYPE,
                    DATA_LENGTH,
                    DATA_PRECISION,
                    DATA_SCALE,
                    NULLABLE,
                    DATA_DEFAULT
                FROM ALL_TAB_COLUMNS
                WHERE TABLE_NAME = :table_name
            """
            
            params = {'table_name': table_name.upper()}
            
            if schema:
                query += " AND OWNER = :owner"
                params['owner'] = schema.upper()
            
            query += " ORDER BY COLUMN_ID"
            
            cursor.execute(query, params)
            columns = cursor.fetchall()
            
            for col in columns:
                structure.append({
                    'column_name': col[0],
                    'data_type': col[1],
                    'data_length': col[2],
                    'data_precision': col[3],
                    'data_scale': col[4],
                    'nullable': col[5],
                    'default_value': col[6]
                })
        
        elif db_type == 'databricks':
            # Databricks query using DESCRIBE TABLE
            full_table = build_full_table_name(table_name, schema, catalog_or_db, db_type)
            query = f"DESCRIBE TABLE {full_table}"
            
            cursor.execute(query)
            columns = cursor.fetchall()
            
            for col in columns:
                # Databricks DESCRIBE returns: col_name, data_type, comment
                if col[0] and not col[0].startswith('#'):  # Skip partition info and comments
                    structure.append({
                        'column_name': col[0],
                        'data_type': col[1],
                        'data_length': None,
                        'data_precision': None,
                        'data_scale': None,
                        'nullable': 'Y',  # Databricks doesn't expose this easily
                        'default_value': None
                    })
        
        elif db_type == 'snowflake':
            # Snowflake query using INFORMATION_SCHEMA
            # Qualify INFORMATION_SCHEMA with database name to avoid "no current database" error
            db_qualifier = catalog_or_db if catalog_or_db else session.get('db_database')
            
            if db_qualifier:
                info_schema = f"{db_qualifier}.INFORMATION_SCHEMA.COLUMNS"
            else:
                info_schema = "INFORMATION_SCHEMA.COLUMNS"
            
            query = f"""
                SELECT 
                    COLUMN_NAME,
                    DATA_TYPE,
                    CHARACTER_MAXIMUM_LENGTH,
                    NUMERIC_PRECISION,
                    NUMERIC_SCALE,
                    IS_NULLABLE,
                    COLUMN_DEFAULT
                FROM {info_schema}
                WHERE TABLE_NAME = %s
            """
            
            params = [table_name.upper()]
            
            if schema:
                query += " AND TABLE_SCHEMA = %s"
                params.append(schema.upper())
            
            if catalog_or_db:
                query += " AND TABLE_CATALOG = %s"
                params.append(catalog_or_db.upper())
            
            query += " ORDER BY ORDINAL_POSITION"
            
            cursor.execute(query, params)
            columns = cursor.fetchall()
            
            for col in columns:
                structure.append({
                    'column_name': col[0],
                    'data_type': col[1],
                    'data_length': col[2],
                    'data_precision': col[3],
                    'data_scale': col[4],
                    'nullable': col[5],
                    'default_value': col[6]
                })
        
        return structure
        
    except Exception as error:
        full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        logger.error(f"Error getting structure for {full_name}: {error}")
        return []

def get_table_indexes(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> List[Dict[str, Any]]:
    """Get table indexes - Multi-database support"""
    try:
        db_type = get_db_type()
        index_list = []
        
        if db_type == 'oracle':
            # Oracle query
            query = """
                SELECT 
                    i.INDEX_NAME,
                    i.INDEX_TYPE,
                    i.UNIQUENESS,
                    LISTAGG(ic.COLUMN_NAME, ', ') WITHIN GROUP (ORDER BY ic.COLUMN_POSITION) as COLUMNS,
                    i.STATUS,
                    i.TABLESPACE_NAME
                FROM ALL_INDEXES i
                LEFT JOIN ALL_IND_COLUMNS ic ON i.INDEX_NAME = ic.INDEX_NAME 
                    AND i.TABLE_NAME = ic.TABLE_NAME
                    AND i.OWNER = ic.INDEX_OWNER
                WHERE i.TABLE_NAME = :table_name
            """
            
            params = {'table_name': table_name.upper()}
            
            if schema:
                query += " AND i.OWNER = :owner"
                params['owner'] = schema.upper()
            
            query += """
                GROUP BY i.INDEX_NAME, i.INDEX_TYPE, i.UNIQUENESS, i.STATUS, i.TABLESPACE_NAME
                ORDER BY i.INDEX_NAME
            """
            
            cursor.execute(query, params)
            indexes = cursor.fetchall()
            
            for idx in indexes:
                index_list.append({
                    'index_name': idx[0],
                    'index_type': idx[1],
                    'uniqueness': idx[2],
                    'columns': idx[3],
                    'status': idx[4],
                    'tablespace': idx[5]
                })
        
        elif db_type == 'databricks':
            # Databricks - using SHOW INDEXES
            try:
                full_table = build_full_table_name(table_name, schema, catalog_or_db, db_type)
                query = f"SHOW INDEXES IN {full_table}"
                cursor.execute(query)
                indexes = cursor.fetchall()
                
                for idx in indexes:
                    # Databricks SHOW INDEXES format may vary
                    index_list.append({
                        'index_name': idx[0] if len(idx) > 0 else 'N/A',
                        'index_type': 'INDEX',
                        'uniqueness': 'N/A',
                        'columns': idx[1] if len(idx) > 1 else 'N/A',
                        'status': 'ACTIVE',
                        'tablespace': 'N/A'
                    })
            except:
                # Databricks may not support indexes or table doesn't have any
                pass
        
        elif db_type == 'snowflake':
            # Snowflake doesn't have traditional indexes, but has clustering keys
            # We can show table constraints instead
            # Qualify INFORMATION_SCHEMA with database name to avoid "no current database" error
            db_qualifier = catalog_or_db if catalog_or_db else session.get('db_database')
            
            if db_qualifier:
                table_constraints = f"{db_qualifier}.INFORMATION_SCHEMA.TABLE_CONSTRAINTS"
            else:
                table_constraints = "INFORMATION_SCHEMA.TABLE_CONSTRAINTS"
            
            query = f"""
                SELECT 
                    CONSTRAINT_NAME,
                    CONSTRAINT_TYPE,
                    '' as COLUMNS
                FROM {table_constraints}
                WHERE TABLE_NAME = %s
            """
            
            params = [table_name.upper()]
            
            if schema:
                query += " AND TABLE_SCHEMA = %s"
                params.append(schema.upper())
            
            if catalog_or_db:
                query += " AND TABLE_CATALOG = %s"
                params.append(catalog_or_db.upper())
            
            try:
                cursor.execute(query, params)
                constraints = cursor.fetchall()
                
                for const in constraints:
                    if const[1] in ['PRIMARY KEY', 'UNIQUE']:
                        index_list.append({
                            'index_name': const[0],
                            'index_type': const[1],
                            'uniqueness': 'UNIQUE' if const[1] in ['PRIMARY KEY', 'UNIQUE'] else 'NONUNIQUE',
                            'columns': 'N/A',
                            'status': 'ACTIVE',
                            'tablespace': 'N/A'
                        })
            except:
                pass
        
        return index_list
        
    except Exception as error:
        full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        logger.error(f"Error getting indexes for {full_name}: {error}")
        return []

def get_table_partitions(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> List[Dict[str, Any]]:
    """Get table partition information (Oracle only - Databricks/Snowflake use different partitioning)"""
    try:
        db_type = get_db_type()
        
        # Partitions are primarily an Oracle feature in this context
        if db_type != 'oracle':
            return []
        
        check_query = """
            SELECT PARTITIONING_TYPE, SUBPARTITIONING_TYPE, PARTITION_COUNT
            FROM ALL_PART_TABLES
            WHERE TABLE_NAME = :table_name
        """
        
        params = {'table_name': table_name.upper()}
        
        if schema:
            check_query += " AND OWNER = :owner"
            params['owner'] = schema.upper()
        
        cursor.execute(check_query, params)
        partition_info = cursor.fetchone()
        
        if not partition_info:
            return []
        
        detail_query = """
            SELECT 
                PARTITION_NAME,
                HIGH_VALUE,
                PARTITION_POSITION,
                TABLESPACE_NAME,
                NUM_ROWS
            FROM ALL_TAB_PARTITIONS
            WHERE TABLE_NAME = :table_name
        """
        
        if schema:
            detail_query += " AND TABLE_OWNER = :owner"
        
        detail_query += " ORDER BY PARTITION_POSITION"
        
        cursor.execute(detail_query, params)
        partitions = cursor.fetchall()
        
        partition_list = [{
            'partitioning_type': partition_info[0],
            'subpartitioning_type': partition_info[1],
            'partition_count': partition_info[2],
            'partitions': []
        }]
        
        for part in partitions:
            partition_list[0]['partitions'].append({
                'partition_name': part[0],
                'high_value': part[1],
                'position': part[2],
                'tablespace': part[3],
                'num_rows': part[4]
            })
        
        return partition_list
    except Exception as error:
        full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        logger.error(f"Error getting partitions for {full_name}: {error}")
        return []

def get_primary_key(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> Dict[str, Any]:
    """Get primary key information - Multi-database support"""
    try:
        db_type = get_db_type()
        
        if db_type == 'oracle':
            # Oracle query
            query = """
                SELECT 
                    c.CONSTRAINT_NAME,
                    LISTAGG(cc.COLUMN_NAME, ', ') WITHIN GROUP (ORDER BY cc.POSITION) as COLUMNS,
                    c.STATUS
                FROM ALL_CONSTRAINTS c
                JOIN ALL_CONS_COLUMNS cc ON c.CONSTRAINT_NAME = cc.CONSTRAINT_NAME 
                    AND c.OWNER = cc.OWNER
                WHERE c.TABLE_NAME = :table_name
                    AND c.CONSTRAINT_TYPE = 'P'
            """
            params = {'table_name': table_name.upper()}
            if schema:
                query += " AND c.OWNER = :owner"
                params['owner'] = schema.upper()
            query += " GROUP BY c.CONSTRAINT_NAME, c.STATUS"
            
            cursor.execute(query, params)
            pk = cursor.fetchone()
            
            if pk:
                return {
                    'constraint_name': pk[0],
                    'columns': pk[1],
                    'status': pk[2]
                }
        
        elif db_type == 'databricks':
            # Databricks - use DESCRIBE TABLE EXTENDED
            try:
                full_table = build_full_table_name(table_name, schema, catalog_or_db, db_type)
                query = f"DESCRIBE TABLE EXTENDED {full_table}"
                cursor.execute(query)
                rows = cursor.fetchall()
                
                # Look for primary key in table details
                for row in rows:
                    if row[0] and 'Primary Key' in str(row[1]):
                        return {
                            'constraint_name': 'PK_' + table_name,
                            'columns': str(row[2]) if len(row) > 2 else 'N/A',
                            'status': 'ENABLED'
                        }
            except:
                pass
        
        elif db_type == 'snowflake':
            # Snowflake query
            # Qualify INFORMATION_SCHEMA with database name to avoid "no current database" error
            db_qualifier = catalog_or_db if catalog_or_db else session.get('db_database')
            
            if db_qualifier:
                table_constraints = f"{db_qualifier}.INFORMATION_SCHEMA.TABLE_CONSTRAINTS"
                key_column_usage = f"{db_qualifier}.INFORMATION_SCHEMA.KEY_COLUMN_USAGE"
            else:
                table_constraints = "INFORMATION_SCHEMA.TABLE_CONSTRAINTS"
                key_column_usage = "INFORMATION_SCHEMA.KEY_COLUMN_USAGE"
            
            query = f"""
                SELECT 
                    CONSTRAINT_NAME,
                    '' as COLUMNS,
                    'ENABLED' as STATUS
                FROM {table_constraints}
                WHERE TABLE_NAME = %s
                    AND CONSTRAINT_TYPE = 'PRIMARY KEY'
            """
            params = [table_name.upper()]
            if schema:
                query += " AND TABLE_SCHEMA = %s"
                params.append(schema.upper())
            
            if catalog_or_db:
                query += " AND TABLE_CATALOG = %s"
                params.append(catalog_or_db.upper())
            
            cursor.execute(query, params)
            pk = cursor.fetchone()
            
            if pk:
                # Get columns for this PK
                col_query = f"""
                    SELECT COLUMN_NAME
                    FROM {key_column_usage}
                    WHERE CONSTRAINT_NAME = %s
                """
                col_params = [pk[0]]
                
                # Add schema filter if provided
                if schema:
                    col_query += " AND TABLE_SCHEMA = %s"
                    col_params.append(schema.upper())
                
                # Add catalog/database filter if provided
                if catalog_or_db:
                    col_query += " AND TABLE_CATALOG = %s"
                    col_params.append(catalog_or_db.upper())
                
                col_query += " ORDER BY ORDINAL_POSITION"
                
                cursor.execute(col_query, col_params)
                cols = cursor.fetchall()
                col_list = ', '.join([c[0] for c in cols])
                
                return {
                    'constraint_name': pk[0],
                    'columns': col_list,
                    'status': 'ENABLED'
                }
        
        return {}
        
    except Exception as error:
        full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        logger.error(f"Error getting primary key for {full_name}: {error}")
        return {}

def get_foreign_keys(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> List[Dict[str, Any]]:
    """Get foreign key information - Multi-database support"""
    try:
        db_type = get_db_type()
        fk_list = []
        
        if db_type == 'oracle':
            # Oracle query
            query = """
                SELECT 
                    c.CONSTRAINT_NAME,
                    LISTAGG(cc.COLUMN_NAME, ', ') WITHIN GROUP (ORDER BY cc.POSITION) as COLUMNS,
                    c.R_CONSTRAINT_NAME,
                    rc.TABLE_NAME as REFERENCED_TABLE,
                    c.DELETE_RULE,
                    c.STATUS
                FROM ALL_CONSTRAINTS c
                JOIN ALL_CONS_COLUMNS cc ON c.CONSTRAINT_NAME = cc.CONSTRAINT_NAME 
                    AND c.OWNER = cc.OWNER
                LEFT JOIN ALL_CONSTRAINTS rc ON c.R_CONSTRAINT_NAME = rc.CONSTRAINT_NAME 
                    AND c.R_OWNER = rc.OWNER
                WHERE c.TABLE_NAME = :table_name
                    AND c.CONSTRAINT_TYPE = 'R'
            """
            params = {'table_name': table_name.upper()}
            if schema:
                query += " AND c.OWNER = :owner"
                params['owner'] = schema.upper()
            query += " GROUP BY c.CONSTRAINT_NAME, c.R_CONSTRAINT_NAME, rc.TABLE_NAME, c.DELETE_RULE, c.STATUS"
            
            cursor.execute(query, params)
            fks = cursor.fetchall()
            
            for fk in fks:
                fk_list.append({
                    'constraint_name': fk[0],
                    'columns': fk[1],
                    'referenced_constraint': fk[2],
                    'referenced_table': fk[3],
                    'delete_rule': fk[4],
                    'status': fk[5]
                })
        
        elif db_type == 'databricks':
            # Databricks doesn't have traditional foreign keys
            # Return empty list
            pass
        
        elif db_type == 'snowflake':
            # Snowflake query for foreign keys
            # Qualify INFORMATION_SCHEMA with database name to avoid "no current database" error
            db_qualifier = catalog_or_db if catalog_or_db else session.get('db_database')
            
            if db_qualifier:
                table_constraints = f"{db_qualifier}.INFORMATION_SCHEMA.TABLE_CONSTRAINTS"
                referential_constraints = f"{db_qualifier}.INFORMATION_SCHEMA.REFERENTIAL_CONSTRAINTS"
            else:
                table_constraints = "INFORMATION_SCHEMA.TABLE_CONSTRAINTS"
                referential_constraints = "INFORMATION_SCHEMA.REFERENTIAL_CONSTRAINTS"
            
            query = f"""
                SELECT 
                    TC.CONSTRAINT_NAME,
                    '' as COLUMNS,
                    TC.CONSTRAINT_NAME as R_CONSTRAINT_NAME,
                    RC.TABLE_NAME as REFERENCED_TABLE,
                    'NO ACTION' as DELETE_RULE,
                    'ENABLED' as STATUS
                FROM {table_constraints} TC
                LEFT JOIN {referential_constraints} RC 
                    ON TC.CONSTRAINT_NAME = RC.CONSTRAINT_NAME
                WHERE TC.TABLE_NAME = %s
                    AND TC.CONSTRAINT_TYPE = 'FOREIGN KEY'
            """
            params = [table_name.upper()]
            if schema:
                query += " AND TC.TABLE_SCHEMA = %s"
                params.append(schema.upper())
            
            if catalog_or_db:
                query += " AND TC.TABLE_CATALOG = %s"
                params.append(catalog_or_db.upper())
            
            try:
                cursor.execute(query, params)
                fks = cursor.fetchall()
                
                for fk in fks:
                    fk_list.append({
                        'constraint_name': fk[0],
                        'columns': 'N/A',
                        'referenced_constraint': fk[2],
                        'referenced_table': fk[3],
                        'delete_rule': fk[4],
                        'status': fk[5]
                    })
            except:
                pass
        
        return fk_list
        
    except Exception as error:
        full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        logger.error(f"Error getting foreign keys for {full_name}: {error}")
        return []
    
def get_last_analyzed(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> Dict[str, Any]:
    """Get last analyzed time and statistics - Multi-database support"""
    try:
        db_type = get_db_type()
        
        if db_type == 'oracle':
            # Oracle query
            query = """
                SELECT 
                    LAST_ANALYZED,
                    NUM_ROWS,
                    BLOCKS,
                    AVG_ROW_LEN,
                    SAMPLE_SIZE
                FROM ALL_TABLES
                WHERE TABLE_NAME = :table_name
            """
            params = {'table_name': table_name.upper()}
            if schema:
                query += " AND OWNER = :owner"
                params['owner'] = schema.upper()
            
            cursor.execute(query, params)
            stats = cursor.fetchone()
            
            if stats:
                return {
                    'last_analyzed': stats[0].strftime('%Y-%m-%d %H:%M:%S') if stats[0] else None,
                    'num_rows': stats[1],
                    'blocks': stats[2],
                    'avg_row_len': stats[3],
                    'sample_size': stats[4],
                    'stale_stats': 'NO'
                }
        
        elif db_type == 'databricks':
            # Databricks - use table statistics
            try:
                full_table = build_full_table_name(table_name, schema, catalog_or_db, db_type)
                query = f"DESCRIBE DETAIL {full_table}"
                cursor.execute(query)
                details = cursor.fetchone()
                
                if details:
                    return {
                        'last_analyzed': 'N/A',
                        'num_rows': details[7] if len(details) > 7 else 'N/A',
                        'blocks': 'N/A',
                        'avg_row_len': 'N/A',
                        'sample_size': 'N/A',
                        'stale_stats': 'N/A'
                    }
            except:
                pass
        
        elif db_type == 'snowflake':
            # Snowflake doesn't expose last analyzed easily
            # Return minimal info
            return {
                'last_analyzed': 'N/A',
                'num_rows': 'N/A',
                'blocks': 'N/A',
                'avg_row_len': 'N/A',
                'sample_size': 'N/A',
                'stale_stats': 'N/A'
            }
        
        return {}
        
    except Exception as error:
        full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        logger.error(f"Error getting last analyzed info for {full_name}: {error}")
        return {}

def get_table_grants(cursor, table_name: str, schema: str = None, catalog_or_db: str = None) -> List[Dict[str, Any]]:
    """Get table grants/privileges - Multi-database support"""
    try:
        db_type = get_db_type()
        grant_list = []
        
        if db_type == 'oracle':
            # Oracle query
            query = """
                SELECT 
                    GRANTEE,
                    PRIVILEGE,
                    GRANTABLE,
                    GRANTOR
                FROM ALL_TAB_PRIVS
                WHERE TABLE_NAME = :table_name
            """
            params = {'table_name': table_name.upper()}
            if schema:
                query += " AND TABLE_SCHEMA = :owner"
                params['owner'] = schema.upper()
            query += " ORDER BY GRANTEE, PRIVILEGE"
            
            cursor.execute(query, params)
            grants = cursor.fetchall()
            
            for grant in grants:
                grant_list.append({
                    'grantee': grant[0],
                    'privilege': grant[1],
                    'grantable': grant[2],
                    'grantor': grant[3]
                })
        
        elif db_type == 'databricks':
            # Databricks - use SHOW GRANTS
            try:
                full_table = build_full_table_name(table_name, schema, catalog_or_db, db_type)
                query = f"SHOW GRANTS ON TABLE {full_table}"
                cursor.execute(query)
                grants = cursor.fetchall()
                
                for grant in grants:
                    grant_list.append({
                        'grantee': grant[0] if len(grant) > 0 else 'N/A',
                        'privilege': grant[1] if len(grant) > 1 else 'N/A',
                        'grantable': 'N/A',
                        'grantor': 'N/A'
                    })
            except:
                pass
        
        elif db_type == 'snowflake':
            # Snowflake grants - use SHOW GRANTS
            try:
                full_table = build_full_table_name(table_name, schema, catalog_or_db, db_type)
                query = f"SHOW GRANTS ON TABLE {full_table}"
                cursor.execute(query)
                grants = cursor.fetchall()
                
                for grant in grants:
                    grant_list.append({
                        'grantee': grant[1] if len(grant) > 1 else 'N/A',
                        'privilege': grant[2] if len(grant) > 2 else 'N/A',
                        'grantable': 'N/A',
                        'grantor': grant[4] if len(grant) > 4 else 'N/A'
                    })
            except:
                pass
        
        return grant_list
        
    except Exception as error:
        full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
        logger.error(f"Error getting grants for {full_name}: {error}")
        return []

def create_excel_report(data: List[Dict[str, Any]]) -> io.BytesIO:
    """Create Excel report with each table's complete information in a single tab"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    section_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Create a sheet for each table with all information
    for table in data:
        # Create sheet name with schema if available
        owner = table.get('count', {}).get('owner')
        table_name = table['table_name']
        
        if owner:
            sheet_name = f"{owner}.{table_name}"[:31]  # Excel sheet name limit
        else:
            sheet_name = table_name[:31]
        
        ws = wb.create_sheet(sheet_name)
        
        # Title
        full_table_name = f"{owner}.{table_name}" if owner else table_name
        ws.append([f"Table Analysis: {full_table_name}"])
        ws.merge_cells(f'A1:G1')
        ws['A1'].font = Font(bold=True, size=14, color="4472C4")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws.append([])
        
        # BASIC INFORMATION SECTION
        ws.append(["BASIC INFORMATION"])
        ws[ws.max_row][0].fill = section_fill
        ws[ws.max_row][0].font = section_font
        ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
        ws.append([])
        
        ws.append(["Metric", "Value"])
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        ws.append(["Schema/Owner", owner if owner else 'N/A'])
        ws.append(["Table Name", table['table_name']])
        ws.append(["Full Name", full_table_name])
        ws.append(["Row Count", table['count'].get('row_count', 'N/A')])
        ws.append(["Column Count", len(table.get('structure', []))])
        ws.append(["Index Count", len(table.get('indexes', []))])
        ws.append(["Partition Count", len(table.get('partitions', []))])
        ws.append(["Has Primary Key", 'Yes' if table.get('primary_key', {}).get('constraint_name') else 'No'])
        ws.append(["Foreign Key Count", len(table.get('foreign_keys', []))])
        ws.append([])
        
        # STATISTICS SECTION
        stats = table.get('last_analyzed', {})
        if stats and any(stats.values()):  # Only show if there's actual data
            ws.append(["STATISTICS"])
            ws[ws.max_row][0].fill = section_fill
            ws[ws.max_row][0].font = section_font
            ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
            ws.append([])
            
            ws.append(["Metric", "Value"])
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            ws.append(["Last Analyzed", str(stats.get('last_analyzed', 'Never'))])
            ws.append(["Num Rows (Stats)", str(stats.get('num_rows', 'N/A'))])
            ws.append(["Blocks", str(stats.get('blocks', 'N/A'))])
            ws.append(["Avg Row Length", str(stats.get('avg_row_len', 'N/A'))])
            ws.append(["Sample Size", str(stats.get('sample_size', 'N/A'))])
            ws.append(["Stale Stats", str(stats.get('stale_stats', 'N/A'))])
            ws.append([])
        
        # COLUMN STRUCTURE SECTION
        if table.get('structure'):
            ws.append(["COLUMN STRUCTURE"])
            ws[ws.max_row][0].fill = section_fill
            ws[ws.max_row][0].font = section_font
            ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
            ws.append([])
            
            ws.append(["Column Name", "Data Type", "Length", "Precision", "Scale", "Nullable", "Default Value"])
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for col in table['structure']:
                ws.append([
                    col.get('column_name'),
                    col.get('data_type'),
                    col.get('data_length'),
                    col.get('data_precision'),
                    col.get('data_scale'),
                    col.get('nullable'),
                    str(col.get('default_value', ''))[:100] if col.get('default_value') else ''
                ])
            ws.append([])
        
        # PRIMARY KEY SECTION
        ws.append(["PRIMARY KEY"])
        ws[ws.max_row][0].fill = section_fill
        ws[ws.max_row][0].font = section_font
        ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
        ws.append([])
        
        if table.get('primary_key', {}).get('constraint_name'):
            ws.append(["Constraint Name", table['primary_key'].get('constraint_name')])
            ws.append(["Columns", table['primary_key'].get('columns')])
            ws.append(["Status", table['primary_key'].get('status')])
        else:
            ws.append(["No primary key defined"])
        ws.append([])
        
        # FOREIGN KEYS SECTION
        ws.append(["FOREIGN KEYS"])
        ws[ws.max_row][0].fill = section_fill
        ws[ws.max_row][0].font = section_font
        ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
        ws.append([])
        
        if table.get('foreign_keys'):
            ws.append(["Constraint Name", "Columns", "Referenced Table", "Referenced Constraint", "Delete Rule", "Status"])
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for fk in table['foreign_keys']:
                ws.append([
                    fk.get('constraint_name'),
                    fk.get('columns'),
                    fk.get('referenced_table'),
                    fk.get('referenced_constraint'),
                    fk.get('delete_rule'),
                    fk.get('status')
                ])
        else:
            ws.append(["No foreign keys defined"])
        ws.append([])
        
        # INDEXES SECTION
        if table.get('indexes'):
            ws.append(["INDEXES"])
            ws[ws.max_row][0].fill = section_fill
            ws[ws.max_row][0].font = section_font
            ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
            ws.append([])
            
            ws.append(["Index Name", "Type", "Uniqueness", "Columns", "Status", "Tablespace"])
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for idx in table['indexes']:
                ws.append([
                    idx.get('index_name'),
                    idx.get('index_type'),
                    idx.get('uniqueness'),
                    idx.get('columns'),
                    idx.get('status'),
                    idx.get('tablespace')
                ])
            ws.append([])
        
        # PARTITIONS SECTION
        if table.get('partitions'):
            ws.append(["PARTITIONS"])
            ws[ws.max_row][0].fill = section_fill
            ws[ws.max_row][0].font = section_font
            ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
            ws.append([])
            
            part_info = table['partitions'][0]
            ws.append(["Partitioning Type", part_info.get('partitioning_type')])
            ws.append(["Subpartitioning Type", part_info.get('subpartitioning_type', 'None')])
            ws.append(["Partition Count", part_info.get('partition_count')])
            ws.append([])
            
            if part_info.get('partitions'):
                ws.append(["Partition Name", "Position", "High Value", "Tablespace", "Num Rows"])
                for cell in ws[ws.max_row]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                
                for part in part_info['partitions']:
                    ws.append([
                        part.get('partition_name'),
                        part.get('position'),
                        str(part.get('high_value', ''))[:100],
                        part.get('tablespace'),
                        part.get('num_rows')
                    ])
            ws.append([])
        
        # GRANTS SECTION
        if table.get('grants'):
            ws.append(["GRANTS"])
            ws[ws.max_row][0].fill = section_fill
            ws[ws.max_row][0].font = section_font
            ws.merge_cells(f'A{ws.max_row}:G{ws.max_row}')
            ws.append([])
            
            ws.append(["Grantee", "Privilege", "Grantable", "Grantor"])
            for cell in ws[ws.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for grant in table['grants']:
                ws.append([
                    grant.get('grantee'),
                    grant.get('privilege'),
                    grant.get('grantable'),
                    grant.get('grantor')
                ])
            ws.append([])
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 12), 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output

def parse_excel_file(file_path):
    """Parse Excel file and extract table data"""
    wb = load_workbook(file_path, data_only=True)
    tables_data = {}
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        table_data = {}
        current_section = None
        section_data = []
        
        for row in ws.iter_rows(values_only=True):
            if not any(row):  # Skip empty rows
                continue
            
            # Detect sections
            if row[0] and isinstance(row[0], str):
                if row[0] == "BASIC INFORMATION":
                    current_section = "basic_info"
                    section_data = []
                elif row[0] == "STATISTICS":
                    if current_section == "basic_info":
                        table_data['basic_info'] = section_data
                    current_section = "statistics"
                    section_data = []
                elif row[0] == "COLUMN STRUCTURE":
                    if current_section == "statistics":
                        table_data['statistics'] = section_data
                    current_section = "columns"
                    section_data = []
                elif row[0] == "PRIMARY KEY":
                    if current_section == "columns":
                        table_data['columns'] = section_data
                    current_section = "primary_key"
                    section_data = []
                elif row[0] == "FOREIGN KEYS":
                    if current_section == "primary_key":
                        table_data['primary_key'] = section_data
                    current_section = "foreign_keys"
                    section_data = []
                elif row[0] == "INDEXES":
                    if current_section == "foreign_keys":
                        table_data['foreign_keys'] = section_data
                    current_section = "indexes"
                    section_data = []
                elif row[0] == "PARTITIONS":
                    if current_section == "indexes":
                        table_data['indexes'] = section_data
                    current_section = "partitions"
                    section_data = []
                elif row[0] == "GRANTS":
                    if current_section == "partitions":
                        table_data['partitions'] = section_data
                    current_section = "grants"
                    section_data = []
                elif current_section and row[0] not in ["Metric", "Column Name", "Constraint Name", "Index Name", "Partition Name", "Grantee", "Partitioning Type"]:
                    section_data.append(row)
        
        # Save last section
        if current_section == "grants":
            table_data['grants'] = section_data
        
        tables_data[sheet_name] = table_data
    
    wb.close()
    return tables_data

def compare_excel_files(file1_path, file2_path):
    """Compare two Excel files and generate differences report"""
    try:
        # Parse both files
        data1 = parse_excel_file(file1_path)
        data2 = parse_excel_file(file2_path)
        
        differences = []
        all_sheets = set(list(data1.keys()) + list(data2.keys()))
        
        for sheet_name in all_sheets:
            sheet_diffs = {
                'sheet_name': sheet_name,
                'differences': []
            }
            
            # Check if sheet exists in both files
            if sheet_name not in data1:
                sheet_diffs['differences'].append({
                    'type': 'SHEET_MISSING',
                    'location': 'File 1',
                    'message': f"Sheet '{sheet_name}' exists only in File 2"
                })
                differences.append(sheet_diffs)
                continue
            
            if sheet_name not in data2:
                sheet_diffs['differences'].append({
                    'type': 'SHEET_MISSING',
                    'location': 'File 2',
                    'message': f"Sheet '{sheet_name}' exists only in File 1"
                })
                differences.append(sheet_diffs)
                continue
            
            # Compare sections
            table1 = data1[sheet_name]
            table2 = data2[sheet_name]
            
            # Compare basic info
            if 'basic_info' in table1 and 'basic_info' in table2:
                for i, (row1, row2) in enumerate(zip(table1['basic_info'], table2['basic_info'])):
                    if row1 != row2:
                        sheet_diffs['differences'].append({
                            'type': 'BASIC_INFO_DIFF',
                            'section': 'Basic Information',
                            'metric': row1[0] if row1 else 'Unknown',
                            'file1_value': row1[1] if len(row1) > 1 else None,
                            'file2_value': row2[1] if len(row2) > 1 else None
                        })
            
            # Compare columns
            if 'columns' in table1 and 'columns' in table2:
                cols1 = {row[0]: row for row in table1['columns'] if row and row[0] not in ['Column Name']}
                cols2 = {row[0]: row for row in table2['columns'] if row and row[0] not in ['Column Name']}
                
                all_cols = set(list(cols1.keys()) + list(cols2.keys()))
                for col_name in all_cols:
                    if col_name not in cols1:
                        sheet_diffs['differences'].append({
                            'type': 'COLUMN_ADDED',
                            'section': 'Column Structure',
                            'column': col_name,
                            'message': f"Column '{col_name}' added in File 2"
                        })
                    elif col_name not in cols2:
                        sheet_diffs['differences'].append({
                            'type': 'COLUMN_REMOVED',
                            'section': 'Column Structure',
                            'column': col_name,
                            'message': f"Column '{col_name}' removed in File 2"
                        })
                    elif cols1[col_name] != cols2[col_name]:
                        sheet_diffs['differences'].append({
                            'type': 'COLUMN_MODIFIED',
                            'section': 'Column Structure',
                            'column': col_name,
                            'file1_value': cols1[col_name],
                            'file2_value': cols2[col_name]
                        })
            
            # Compare indexes
            if 'indexes' in table1 and 'indexes' in table2:
                idx1 = {row[0]: row for row in table1['indexes'] if row and row[0] not in ['Index Name']}
                idx2 = {row[0]: row for row in table2['indexes'] if row and row[0] not in ['Index Name']}
                
                all_idx = set(list(idx1.keys()) + list(idx2.keys()))
                for idx_name in all_idx:
                    if idx_name not in idx1:
                        sheet_diffs['differences'].append({
                            'type': 'INDEX_ADDED',
                            'section': 'Indexes',
                            'index': idx_name,
                            'message': f"Index '{idx_name}' added in File 2"
                        })
                    elif idx_name not in idx2:
                        sheet_diffs['differences'].append({
                            'type': 'INDEX_REMOVED',
                            'section': 'Indexes',
                            'index': idx_name,
                            'message': f"Index '{idx_name}' removed in File 2"
                        })
                    elif idx1[idx_name] != idx2[idx_name]:
                        sheet_diffs['differences'].append({
                            'type': 'INDEX_MODIFIED',
                            'section': 'Indexes',
                            'index': idx_name,
                            'file1_value': idx1[idx_name],
                            'file2_value': idx2[idx_name]
                        })
            
            if sheet_diffs['differences']:
                differences.append(sheet_diffs)
        
        return differences
    except Exception as e:
        logger.error(f"Error comparing files: {str(e)}")
        raise

def create_comparison_report(differences):
    """Create Excel report showing differences"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.append(["Comparison Summary"])
    ws_summary.merge_cells('A1:D1')
    ws_summary['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws_summary['A1'].alignment = Alignment(horizontal="center")
    ws_summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append([])
    
    ws_summary.append(["Sheet/Table", "Total Differences", "Added", "Removed", "Modified"])
    for cell in ws_summary[ws_summary.max_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
    
    total_diffs = 0
    for table_diff in differences:
        sheet_name = table_diff['sheet_name']
        diffs = table_diff['differences']
        
        added = len([d for d in diffs if 'ADDED' in d['type']])
        removed = len([d for d in diffs if 'REMOVED' in d['type']])
        modified = len([d for d in diffs if 'MODIFIED' in d['type'] or 'DIFF' in d['type']])
        
        ws_summary.append([sheet_name, len(diffs), added, removed, modified])
        total_diffs += len(diffs)
    
    ws_summary.append([])
    ws_summary.append(["Total Differences:", total_diffs])
    ws_summary['A' + str(ws_summary.max_row)].font = Font(bold=True)
    
    # Auto-adjust column widths
    for column in ws_summary.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws_summary.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 50)
    
    # Detailed differences for each table
    for table_diff in differences:
        sheet_name = table_diff['sheet_name'][:31]
        ws = wb.create_sheet(sheet_name)
        
        ws.append([f"Differences for: {table_diff['sheet_name']}"])
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(bold=True, size=14, color="4472C4")
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.append([])
        
        ws.append(["Type", "Section", "Item", "File 1 Value", "File 2 Value"])
        for cell in ws[ws.max_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
        
        
        for diff in table_diff['differences']:
            diff_type = diff.get('type', 'UNKNOWN')
            section = diff.get('section', 'N/A')
            
            # Determine item name based on type
            if 'column' in diff:
                item = diff['column']
            elif 'index' in diff:
                item = diff['index']
            elif 'metric' in diff:
                item = diff['metric']
            elif 'message' in diff:
                item = diff['message']
            else:
                item = 'N/A'
            
            file1_val = str(diff.get('file1_value', 'N/A'))[:100]
            file2_val = str(diff.get('file2_value', 'N/A'))[:100]
            
            row = ws.append([diff_type, section, item, file1_val, file2_val])
            
            # Apply color coding based on type
            row_num = ws.max_row
            if 'ADDED' in diff_type:
                for col in range(1, 6):
                    ws.cell(row_num, col).fill = added_fill
            elif 'REMOVED' in diff_type:
                for col in range(1, 6):
                    ws.cell(row_num, col).fill = removed_fill
            elif 'MODIFIED' in diff_type or 'DIFF' in diff_type:
                for col in range(1, 6):
                    ws.cell(row_num, col).fill = modified_fill
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 15), 50)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


@app.route('/')
def home():
    """Root endpoint - show home page"""
    return render_template('home.html')

@app.route('/login')
def login_page():
    """Serve the login page"""
    return render_template('login.html')

@app.route('/query-optimizer')
def query_optimizer_page():
    """Serve the SQL Query Optimizer page"""
    return render_template('query-optimizer.html')

@app.route('/nl-to-sql')
def nl_to_sql_page():
    """Serve the Natural Language to SQL Generator page"""
    return render_template('nl-to-sql.html')

@app.route('/sql-script-generator')
def sql_script_generator_page():
    """Serve the SQL Script Generator page"""
    return render_template('sql-script-generator.html')

@app.route('/api/login', methods=['POST'])
def login():
    """Login endpoint - validates credentials and creates session for multiple database types"""
    try:
        data = request.get_json()
        db_type = data.get('db_type', 'oracle').lower()
        
        # Oracle Login
        if db_type == 'oracle':
            username = data.get('username')
            password = data.get('password')
            dsn = data.get('dsn')
            
            if not username or not password or not dsn:
                return jsonify({
                    'success': False,
                    'error': 'Username, password, and DSN are required for Oracle'
                }), 400
            
            try:
                connection = oracledb.connect(
                    user=username,
                    password=password,
                    dsn=dsn
                )
                connection.close()
                
                # Store credentials in session
                session.permanent = True  # Enable session timeout
                session['db_type'] = 'oracle'
                session['db_user'] = username
                session['db_password'] = password
                session['db_dsn'] = dsn
                session['logged_in'] = True
                
                logger.info(f"Oracle user {username} logged in successfully")
                
                return jsonify({
                    'success': True,
                    'message': 'Oracle login successful',
                    'db_type': 'oracle'
                })
                
            except oracledb.Error as error:
                logger.error(f"Oracle login failed for user {username}: {error}")
                return jsonify({
                    'success': False,
                    'error': 'Invalid Oracle credentials or DSN. Please check your inputs.'
                }), 401
        
        # Databricks Login
        elif db_type == 'databricks':
            if not DATABRICKS_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Databricks connector not installed. Install with: pip install databricks-sql-connector'
                }), 400
            
            server_hostname = data.get('server_hostname')
            http_path = data.get('http_path')
            access_token = data.get('access_token')
            username = data.get('username')
            password = data.get('password')
            authenticator = data.get('authenticator')  # 'azuread' for Azure AD
            
            # Validate required fields
            if not server_hostname or not http_path:
                return jsonify({
                    'success': False,
                    'error': 'Server hostname and HTTP path are required for Databricks'
                }), 400
            
            # Check authentication method
            use_azure_ad = authenticator == 'azuread'
            has_token = access_token and access_token.strip()
            has_credentials = username and password
            
            if not use_azure_ad and not has_token and not has_credentials:
                return jsonify({
                    'success': False,
                    'error': 'Provide Access Token, Username/Password, or use Azure AD authentication'
                }), 400
            
            try:
                # Try to connect based on authentication method
                if use_azure_ad:
                    # Azure AD authentication with headless mode
                    import uuid
                    import os
                    # Suppress OAuth redirect success page
                    os.environ['DATABRICKS_CLI_DO_NOT_REDIRECT'] = '1'
                    
                    connection = databricks_sql.connect(
                        server_hostname=server_hostname,
                        http_path=http_path,
                        auth_type='azure-ad',
                        # Use localhost callback to prevent browser tab message
                        redirect_port=8020,
                        # Experimental: disable browser redirect page
                        experimental_oauth_client_id=None
                    )
                    auth_method = 'Azure AD'
                    
                    # Clean up environment variable
                    if 'DATABRICKS_CLI_DO_NOT_REDIRECT' in os.environ:
                        del os.environ['DATABRICKS_CLI_DO_NOT_REDIRECT']
                    
                    # Cache Azure AD connection to prevent re-authentication
                    session_id = str(uuid.uuid4())
                    session['session_id'] = session_id
                    _connection_cache[session_id] = connection
                    # Don't close Azure AD connection
                elif has_token:
                    # Token authentication
                    connection = databricks_sql.connect(
                        server_hostname=server_hostname,
                        http_path=http_path,
                        access_token=access_token
                    )
                    auth_method = 'token'
                    connection.close()
                else:
                    # Username/password authentication
                    connection = databricks_sql.connect(
                        server_hostname=server_hostname,
                        http_path=http_path,
                        username=username,
                        password=password
                    )
                    auth_method = 'username/password'
                    connection.close()
                
                # Store credentials in session
                session.permanent = True  # Enable session timeout
                session['db_type'] = 'databricks'
                session['db_server_hostname'] = server_hostname
                session['db_http_path'] = http_path
                
                if use_azure_ad:
                    session['db_authenticator'] = 'azuread'
                elif has_token:
                    session['db_access_token'] = access_token
                else:
                    session['db_user'] = username
                    session['db_password'] = password
                
                session['logged_in'] = True
                
                logger.info(f"Databricks user logged in successfully to {server_hostname} using {auth_method}")
                
                return jsonify({
                    'success': True,
                    'message': f'Databricks login successful ({auth_method})',
                    'db_type': 'databricks'
                })
                
            except Exception as error:
                logger.error(f"Databricks login failed: {error}")
                return jsonify({
                    'success': False,
                    'error': f'Invalid Databricks credentials. Error: {str(error)}'
                }), 401
        
        # Snowflake Login
        elif db_type == 'snowflake':
            if not SNOWFLAKE_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Snowflake connector not installed. Install with: pip install snowflake-connector-python'
                }), 400
            
            username = data.get('username')
            password = data.get('password')
            account = data.get('account')
            warehouse = data.get('warehouse')
            database = data.get('database')
            schema = data.get('schema')
            authenticator = data.get('authenticator', 'snowflake')  # Default to standard auth
            
            if not username or not account:
                return jsonify({
                    'success': False,
                    'error': 'Username and account are required for Snowflake'
                }), 400
            
            # Check authentication method
            use_sso = authenticator == 'externalbrowser'
            
            if not use_sso and not password:
                return jsonify({
                    'success': False,
                    'error': 'Password is required for standard authentication. Check "Use SSO" if you login with SSO.'
                }), 400
            
            try:
                # Build connection parameters
                conn_params = {
                    'user': username,
                    'account': account
                }
                
                if warehouse:
                    conn_params['warehouse'] = warehouse
                if database:
                    conn_params['database'] = database
                if schema:
                    conn_params['schema'] = schema
                
                # Add authentication
                if use_sso:
                    conn_params['authenticator'] = 'externalbrowser'
                    conn_params['login_timeout'] = 120  # 2 minutes for browser authentication
                    auth_method = 'SSO (Browser)'
                    logger.info(f"Attempting Snowflake SSO login for {username} - Browser window should open...")
                else:
                    conn_params['password'] = password
                    auth_method = 'password'
                
                connection = snowflake.connector.connect(**conn_params)
                
                # Cache SSO connection to prevent re-authentication
                if use_sso:
                    import uuid
                    session_id = str(uuid.uuid4())
                    session['session_id'] = session_id
                    _connection_cache[session_id] = connection
                    # Don't close SSO connection
                else:
                    connection.close()
                
                # Store credentials in session
                session.permanent = True  # Enable session timeout
                session['db_type'] = 'snowflake'
                session['db_user'] = username
                session['db_account'] = account
                session['db_warehouse'] = warehouse
                session['db_database'] = database
                session['db_schema'] = schema
                
                if use_sso:
                    session['db_authenticator'] = 'externalbrowser'
                else:
                    session['db_password'] = password
                
                session['logged_in'] = True
                
                logger.info(f"Snowflake user {username} logged in successfully to {account} using {auth_method}")
                
                return jsonify({
                    'success': True,
                    'message': f'Snowflake login successful ({auth_method})',
                    'db_type': 'snowflake'
                })
                
            except Exception as error:
                logger.error(f"Snowflake login failed for user {username}: {error}")
                return jsonify({
                    'success': False,
                    'error': f'Invalid Snowflake credentials. Error: {str(error)}'
                }), 401
        
        else:
            return jsonify({
                'success': False,
                'error': f'Unsupported database type: {db_type}. Supported types: oracle, databricks, snowflake'
            }), 400
            
    except Exception as e:
        logger.error(f"Error in login: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/logout', methods=['POST'])
def logout():
    """Logout endpoint - clears session"""
    username = session.get('db_user', 'Unknown')
    
    # Clear cached connection if exists
    session_id = session.get('session_id')
    if session_id and session_id in _connection_cache:
        try:
            _connection_cache[session_id].close()
        except:
            pass
        del _connection_cache[session_id]
    
    session.clear()
    logger.info(f"User {username} logged out")
    return jsonify({
        'success': True,
        'message': 'Logged out successfully'
    })


# Dual database connections storage (for source-target and SQL query compare)
_dual_connections = {}

@app.route('/api/dual-login', methods=['POST'])
def dual_login():
    """Login endpoint for dual database connections (source and target can be different)"""
    try:
        import uuid
        data = request.get_json()
        db_type = data.get('db_type', 'oracle').lower()
        connection_name = data.get('connection_name', 'source')  # 'source' or 'target'
        
        # Create unique session ID for this connection
        session_id = str(uuid.uuid4())
        
        # Establish connection based on database type
        if db_type == 'oracle':
            username = data.get('username')
            password = data.get('password')
            dsn = data.get('dsn')
            
            if not username or not password or not dsn:
                return jsonify({
                    'success': False,
                    'error': 'Username, password, and DSN are required for Oracle'
                }), 400
            
            try:
                import oracledb
                connection = oracledb.connect(
                    user=username,
                    password=password,
                    dsn=dsn
                )
                
                # Store connection
                _dual_connections[session_id] = {
                    'connection': connection,
                    'db_type': 'oracle',
                    'connection_name': connection_name,
                    'username': username
                }
                
                logger.info(f"Dual Oracle connection established: {connection_name} for {username}")
                
                return jsonify({
                    'success': True,
                    'message': 'Oracle connection successful',
                    'session_id': session_id,
                    'db_type': 'oracle'
                })
                
            except Exception as error:
                logger.error(f"Dual Oracle login failed: {error}")
                return jsonify({
                    'success': False,
                    'error': str(error)
                }), 401
        
        elif db_type == 'databricks':
            if not DATABRICKS_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Databricks connector not installed'
                }), 400
            
            server_hostname = data.get('server_hostname')
            http_path = data.get('http_path')
            access_token = data.get('access_token')
            authenticator = data.get('authenticator')  # 'azuread' for Azure AD
            
            # Validate required fields
            if not server_hostname or not http_path:
                return jsonify({
                    'success': False,
                    'error': 'Server hostname and HTTP path are required for Databricks'
                }), 400
            
            # Check authentication method
            use_azure_ad = authenticator == 'azuread'
            has_token = access_token and access_token.strip()
            
            if not use_azure_ad and not has_token:
                return jsonify({
                    'success': False,
                    'error': 'Provide Access Token or use Azure AD authentication'
                }), 400
            
            try:
                # Connect based on authentication method
                if use_azure_ad:
                    # Azure AD authentication
                    import os
                    os.environ['DATABRICKS_CLI_DO_NOT_REDIRECT'] = '1'
                    
                    connection = databricks_sql.connect(
                        server_hostname=server_hostname,
                        http_path=http_path,
                        auth_type='azure-ad',
                        redirect_port=8020
                    )
                    auth_method = 'Azure AD'
                    
                    # Clean up environment variable
                    if 'DATABRICKS_CLI_DO_NOT_REDIRECT' in os.environ:
                        del os.environ['DATABRICKS_CLI_DO_NOT_REDIRECT']
                else:
                    # Token authentication
                    connection = databricks_sql.connect(
                        server_hostname=server_hostname,
                        http_path=http_path,
                        access_token=access_token
                    )
                    auth_method = 'token'
                
                _dual_connections[session_id] = {
                    'connection': connection,
                    'db_type': 'databricks',
                    'connection_name': connection_name,
                    'server_hostname': server_hostname
                }
                
                logger.info(f"Dual Databricks connection established: {connection_name} using {auth_method}")
                
                return jsonify({
                    'success': True,
                    'message': f'Databricks connection successful ({auth_method})',
                    'session_id': session_id,
                    'db_type': 'databricks'
                })
                
            except Exception as error:
                logger.error(f"Dual Databricks login failed: {error}")
                return jsonify({
                    'success': False,
                    'error': str(error)
                }), 401
        
        elif db_type == 'snowflake':
            if not SNOWFLAKE_AVAILABLE:
                return jsonify({
                    'success': False,
                    'error': 'Snowflake connector not installed'
                }), 400
            
            username = data.get('username')
            password = data.get('password')
            account = data.get('account')
            warehouse = data.get('warehouse')
            database = data.get('database')
            schema = data.get('schema')
            authenticator = data.get('authenticator', 'snowflake')  # Default to standard auth
            
            # Validate required fields
            if not username or not account:
                return jsonify({
                    'success': False,
                    'error': 'Username and account are required for Snowflake'
                }), 400
            
            # Check authentication method
            use_sso = authenticator == 'externalbrowser'
            
            if not use_sso and not password:
                return jsonify({
                    'success': False,
                    'error': 'Password is required for standard authentication. Check "Use SSO" if you login with SSO.'
                }), 400
            
            try:
                # Build connection parameters
                conn_params = {
                    'user': username,
                    'account': account
                }
                
                # Add optional parameters
                if warehouse:
                    conn_params['warehouse'] = warehouse
                if database:
                    conn_params['database'] = database
                if schema:
                    conn_params['schema'] = schema
                
                # Add authentication
                if use_sso:
                    conn_params['authenticator'] = 'externalbrowser'
                    conn_params['login_timeout'] = 120  # 2 minutes for browser authentication
                    auth_method = 'SSO (Browser)'
                    logger.info(f"Attempting Snowflake SSO login for {username} - Browser window should open...")
                else:
                    conn_params['password'] = password
                    auth_method = 'password'
                
                connection = snowflake.connector.connect(**conn_params)
                
                _dual_connections[session_id] = {
                    'connection': connection,
                    'db_type': 'snowflake',
                    'connection_name': connection_name,
                    'username': username
                }
                
                logger.info(f"Dual Snowflake connection established: {connection_name} using {auth_method}")
                
                return jsonify({
                    'success': True,
                    'message': f'Snowflake connection successful ({auth_method})',
                    'session_id': session_id,
                    'db_type': 'snowflake'
                })
                
            except Exception as error:
                logger.error(f"Dual Snowflake login failed: {error}")
                return jsonify({
                    'success': False,
                    'error': str(error)
                }), 401
        
        else:
            return jsonify({
                'success': False,
                'error': f'Unsupported database type: {db_type}'
            }), 400
            
    except Exception as e:
        logger.error(f"Error in dual-login: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/check-auth', methods=['GET'])
def check_auth():
    """Check if user is authenticated"""
    if session.get('logged_in'):
        db_type = session.get('db_type', 'unknown')
        response = {
            'authenticated': True,
            'db_type': db_type
        }
        
        # Add username for Oracle and Snowflake
        if db_type in ['oracle', 'snowflake']:
            response['username'] = session.get('db_user', 'Unknown')
        # Add server hostname for Databricks
        elif db_type == 'databricks':
            response['server'] = session.get('db_server_hostname', 'Unknown')
        
        return jsonify(response)
    
    return jsonify({
        'authenticated': False
    })

@app.route('/api/session-info', methods=['GET'])
def session_info():
    """Get session timeout information"""
    if not session.get('logged_in'):
        return jsonify({
            'authenticated': False,
            'timeout_minutes': 0,
            'time_remaining_seconds': 0
        })
    
    # Calculate session timeout info
    timeout_minutes = app.config['PERMANENT_SESSION_LIFETIME'].total_seconds() / 60
    
    return jsonify({
        'authenticated': True,
        'timeout_minutes': int(timeout_minutes),
        'session_lifetime_seconds': int(app.config['PERMANENT_SESSION_LIFETIME'].total_seconds())
    })

@app.route('/api/keepalive', methods=['POST'])
def keepalive():
    """Keep session alive - refresh session timeout"""
    if session.get('logged_in'):
        # Just accessing session refreshes it when SESSION_REFRESH_EACH_REQUEST is True
        session.modified = True
        return jsonify({
            'success': True,
            'message': 'Session refreshed'
        })
    return jsonify({
        'success': False,
        'message': 'Not authenticated'
    }), 401

@app.route('/api/compare-source-target-dual', methods=['POST'])
def compare_source_target_dual():
    """Compare source and target tables using dual database connections"""
    try:
        data = request.get_json()
        source_session = data.get('source_session')
        target_session = data.get('target_session')
        source_info = data.get('source', {})
        target_info = data.get('target', {})
        options = data.get('options', {})
        
        # Get connections from dual connections storage
        if source_session not in _dual_connections:
            return jsonify({'success': False, 'error': 'Source connection not found or expired'}), 400
        
        if target_session not in _dual_connections:
            return jsonify({'success': False, 'error': 'Target connection not found or expired'}), 400
        
        source_conn = _dual_connections[source_session]['connection']
        target_conn = _dual_connections[target_session]['connection']
        source_db_type = _dual_connections[source_session]['db_type']
        target_db_type = _dual_connections[target_session]['db_type']
        
        source_table = source_info.get('table')
        source_schema = source_info.get('schema')
        target_table = target_info.get('table')
        target_schema = target_info.get('schema')
        
        if not source_table or not target_table:
            return jsonify({'success': False, 'error': 'Table names are required'}), 400
        
        differences = {
            'total_count': 0,
            'structure': [],
            'row_count': {},
            'indexes': [],
            'constraints': []
        }
        
        source_cursor = source_conn.cursor()
        target_cursor = target_conn.cursor()
        
        try:
            # Helper function to get structure using db-specific cursor
            def get_structure_dual(cursor, table, schema, db_type):
                """Get table structure for dual connection"""
                structure = []
                try:
                    if db_type == 'oracle':
                        query = """
                            SELECT 
                                COLUMN_NAME,
                                DATA_TYPE,
                                DATA_LENGTH,
                                DATA_PRECISION,
                                DATA_SCALE,
                                NULLABLE,
                                DATA_DEFAULT
                            FROM ALL_TAB_COLUMNS
                            WHERE TABLE_NAME = :table_name
                        """
                        params = {'table_name': table.upper()}
                        if schema:
                            query += " AND OWNER = :owner"
                            params['owner'] = schema.upper()
                        query += " ORDER BY COLUMN_ID"
                        
                        cursor.execute(query, params)
                        columns = cursor.fetchall()
                        
                        for col in columns:
                            structure.append({
                                'column_name': col[0],
                                'data_type': col[1],
                                'data_length': col[2],
                                'data_precision': col[3],
                                'data_scale': col[4],
                                'nullable': col[5],
                                'default_value': col[6]
                            })
                    
                    elif db_type == 'databricks':
                        full_table = f"{schema}.{table}" if schema else table
                        query = f"DESCRIBE TABLE {full_table}"
                        cursor.execute(query)
                        columns = cursor.fetchall()
                        
                        for col in columns:
                            if col[0] and not col[0].startswith('#'):
                                structure.append({
                                    'column_name': col[0],
                                    'data_type': col[1],
                                    'data_length': None,
                                    'data_precision': None,
                                    'data_scale': None,
                                    'nullable': 'Y',
                                    'default_value': None
                                })
                    
                    elif db_type == 'snowflake':
                        query = """
                            SELECT 
                                COLUMN_NAME,
                                DATA_TYPE,
                                CHARACTER_MAXIMUM_LENGTH,
                                NUMERIC_PRECISION,
                                NUMERIC_SCALE,
                                IS_NULLABLE,
                                COLUMN_DEFAULT
                            FROM INFORMATION_SCHEMA.COLUMNS
                            WHERE TABLE_NAME = %s
                        """
                        params = [table.upper()]
                        if schema:
                            query += " AND TABLE_SCHEMA = %s"
                            params.append(schema.upper())
                        query += " ORDER BY ORDINAL_POSITION"
                        
                        cursor.execute(query, params)
                        columns = cursor.fetchall()
                        
                        for col in columns:
                            structure.append({
                                'column_name': col[0],
                                'data_type': col[1],
                                'data_length': col[2],
                                'data_precision': col[3],
                                'data_scale': col[4],
                                'nullable': col[5],
                                'default_value': col[6]
                            })
                except Exception as e:
                    logger.error(f"Error getting structure: {e}")
                
                return structure
            
            # Compare structure
            if options.get('structure', False):
                source_structure = get_structure_dual(source_cursor, source_table, source_schema, source_db_type)
                target_structure = get_structure_dual(target_cursor, target_table, target_schema, target_db_type)
                
                source_cols = {col['column_name']: col for col in source_structure}
                target_cols = {col['column_name']: col for col in target_structure}
                
                # Find differences
                all_columns = set(list(source_cols.keys()) + list(target_cols.keys()))
                
                for col_name in all_columns:
                    if col_name not in source_cols:
                        differences['structure'].append({
                            'column_name': col_name,
                            'diff_type': 'Missing in Source',
                            'source_value': None,
                            'target_value': f"{target_cols[col_name]['data_type']}({target_cols[col_name].get('data_length', '')})",
                            'type': 'added'
                        })
                        differences['total_count'] += 1
                    elif col_name not in target_cols:
                        differences['structure'].append({
                            'column_name': col_name,
                            'diff_type': 'Missing in Target',
                            'source_value': f"{source_cols[col_name]['data_type']}({source_cols[col_name].get('data_length', '')})",
                            'target_value': None,
                            'type': 'removed'
                        })
                        differences['total_count'] += 1
                    else:
                        # Check if data type or attributes differ
                        source_col = source_cols[col_name]
                        target_col = target_cols[col_name]
                        
                        if (source_col['data_type'] != target_col['data_type'] or 
                            source_col.get('data_length') != target_col.get('data_length') or
                            source_col.get('data_precision') != target_col.get('data_precision') or
                            source_col.get('nullable') != target_col.get('nullable')):
                            
                            differences['structure'].append({
                                'column_name': col_name,
                                'diff_type': 'Modified',
                                'source_value': f"{source_col['data_type']}({source_col.get('data_length', '')}) {source_col.get('nullable', '')}",
                                'target_value': f"{target_col['data_type']}({target_col.get('data_length', '')}) {target_col.get('nullable', '')}",
                                'type': 'modified'
                            })
                            differences['total_count'] += 1
            
            # Compare row count
            if options.get('rowCount', False):
                source_count_query = f"SELECT COUNT(*) FROM {source_schema + '.' if source_schema else ''}{source_table}"
                target_count_query = f"SELECT COUNT(*) FROM {target_schema + '.' if target_schema else ''}{target_table}"
                
                source_cursor.execute(source_count_query)
                source_count = source_cursor.fetchone()[0]
                
                target_cursor.execute(target_count_query)
                target_count = target_cursor.fetchone()[0]
                
                differences['row_count'] = {
                    'source': source_count,
                    'target': target_count,
                    'different': source_count != target_count,
                    'difference': abs(source_count - target_count)
                }
                
                if source_count != target_count:
                    differences['total_count'] += 1
            
            # Compare indexes
            if options.get('indexes', False):
                # Note: This is a simplified comparison - actual implementation would need db-specific queries
                try:
                    source_indexes = []
                    target_indexes = []
                    
                    # For Oracle
                    if source_db_type == 'oracle':
                        query = """
                            SELECT INDEX_NAME, INDEX_TYPE, UNIQUENESS
                            FROM ALL_INDEXES
                            WHERE TABLE_NAME = :table_name
                        """
                        params = {'table_name': source_table.upper()}
                        if source_schema:
                            query += " AND OWNER = :owner"
                            params['owner'] = source_schema.upper()
                        source_cursor.execute(query, params)
                        source_indexes = [{'index_name': row[0], 'index_type': row[1], 'uniqueness': row[2]} 
                                        for row in source_cursor.fetchall()]
                    
                    if target_db_type == 'oracle':
                        query = """
                            SELECT INDEX_NAME, INDEX_TYPE, UNIQUENESS
                            FROM ALL_INDEXES
                            WHERE TABLE_NAME = :table_name
                        """
                        params = {'table_name': target_table.upper()}
                        if target_schema:
                            query += " AND OWNER = :owner"
                            params['owner'] = target_schema.upper()
                        target_cursor.execute(query, params)
                        target_indexes = [{'index_name': row[0], 'index_type': row[1], 'uniqueness': row[2]} 
                                        for row in target_cursor.fetchall()]
                    
                    source_idx_names = {idx['index_name'] for idx in source_indexes}
                    target_idx_names = {idx['index_name'] for idx in target_indexes}
                    
                    # Missing in target
                    for idx_name in source_idx_names - target_idx_names:
                        idx = next((i for i in source_indexes if i['index_name'] == idx_name), None)
                        if idx:
                            differences['indexes'].append({
                                'index_name': idx_name,
                                'diff_type': idx.get('index_type', 'INDEX'),
                                'status': 'missing_in_target'
                            })
                            differences['total_count'] += 1
                    
                    # Missing in source
                    for idx_name in target_idx_names - source_idx_names:
                        idx = next((i for i in target_indexes if i['index_name'] == idx_name), None)
                        if idx:
                            differences['indexes'].append({
                                'index_name': idx_name,
                                'diff_type': idx.get('index_type', 'INDEX'),
                                'status': 'missing_in_source'
                            })
                            differences['total_count'] += 1
                except Exception as e:
                    logger.error(f"Error comparing indexes: {e}")
            
            # Compare constraints
            if options.get('constraints', False):
                try:
                    # Compare primary keys
                    source_pk = None
                    target_pk = None
                    
                    # For Oracle source
                    if source_db_type == 'oracle':
                        query = """
                            SELECT c.CONSTRAINT_NAME,
                                   LISTAGG(cc.COLUMN_NAME, ', ') WITHIN GROUP (ORDER BY cc.POSITION) as COLUMNS
                            FROM ALL_CONSTRAINTS c
                            JOIN ALL_CONS_COLUMNS cc ON c.CONSTRAINT_NAME = cc.CONSTRAINT_NAME AND c.OWNER = cc.OWNER
                            WHERE c.TABLE_NAME = :table_name AND c.CONSTRAINT_TYPE = 'P'
                        """
                        params = {'table_name': source_table.upper()}
                        if source_schema:
                            query += " AND c.OWNER = :owner"
                            params['owner'] = source_schema.upper()
                        query += " GROUP BY c.CONSTRAINT_NAME"
                        
                        source_cursor.execute(query, params)
                        pk_row = source_cursor.fetchone()
                        if pk_row:
                            source_pk = {'constraint_name': pk_row[0], 'columns': pk_row[1]}
                    
                    # For Oracle target
                    if target_db_type == 'oracle':
                        query = """
                            SELECT c.CONSTRAINT_NAME,
                                   LISTAGG(cc.COLUMN_NAME, ', ') WITHIN GROUP (ORDER BY cc.POSITION) as COLUMNS
                            FROM ALL_CONSTRAINTS c
                            JOIN ALL_CONS_COLUMNS cc ON c.CONSTRAINT_NAME = cc.CONSTRAINT_NAME AND c.OWNER = cc.OWNER
                            WHERE c.TABLE_NAME = :table_name AND c.CONSTRAINT_TYPE = 'P'
                        """
                        params = {'table_name': target_table.upper()}
                        if target_schema:
                            query += " AND c.OWNER = :owner"
                            params['owner'] = target_schema.upper()
                        query += " GROUP BY c.CONSTRAINT_NAME"
                        
                        target_cursor.execute(query, params)
                        pk_row = target_cursor.fetchone()
                        if pk_row:
                            target_pk = {'constraint_name': pk_row[0], 'columns': pk_row[1]}
                    
                    # Compare primary keys
                    has_source_pk = bool(source_pk)
                    has_target_pk = bool(target_pk)
                    
                    if has_source_pk and not has_target_pk:
                        differences['constraints'].append({
                            'constraint_name': source_pk['constraint_name'],
                            'constraint_type': 'PRIMARY KEY',
                            'columns': source_pk.get('columns', ''),
                            'source_value': source_pk.get('columns', ''),
                            'target_value': 'None',
                            'status': 'missing_in_target'
                        })
                        differences['total_count'] += 1
                    elif not has_source_pk and has_target_pk:
                        differences['constraints'].append({
                            'constraint_name': target_pk['constraint_name'],
                            'constraint_type': 'PRIMARY KEY',
                            'columns': target_pk.get('columns', ''),
                            'source_value': 'None',
                            'target_value': target_pk.get('columns', ''),
                            'status': 'missing_in_source'
                        })
                        differences['total_count'] += 1
                    elif has_source_pk and has_target_pk:
                        source_pk_cols = source_pk.get('columns', '').strip()
                        target_pk_cols = target_pk.get('columns', '').strip()
                        
                        if source_pk_cols != target_pk_cols:
                            differences['constraints'].append({
                                'constraint_name': f"{source_pk['constraint_name']} / {target_pk['constraint_name']}",
                                'constraint_type': 'PRIMARY KEY',
                                'columns': 'Different columns',
                                'source_value': source_pk_cols,
                                'target_value': target_pk_cols,
                                'status': 'modified'
                            })
                            differences['total_count'] += 1
                except Exception as e:
                    logger.error(f"Error comparing constraints: {e}")
            
            return jsonify({
                'success': True,
                'source_table': f"{source_schema}.{source_table}" if source_schema else source_table,
                'target_table': f"{target_schema}.{target_table}" if target_schema else target_table,
                'differences': differences
            })
            
        finally:
            source_cursor.close()
            target_cursor.close()
            
    except Exception as e:
        logger.error(f"Error in compare-source-target-dual: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export-comparison', methods=['POST'])
def export_comparison():
    """Export table comparison results to Excel"""
    try:
        comp_data = request.get_json()
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)
        
        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        section_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        section_font = Font(bold=True, color="FFFFFF", size=11)
        added_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        removed_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        modified_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Table Comparison Report"])
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'].font = Font(bold=True, size=14, color="4472C4")
        ws_summary['A1'].alignment = Alignment(horizontal="center")
        ws_summary.append(["Generated:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_summary.append([])
        
        ws_summary.append(["Source Table:", comp_data.get('source_table', 'N/A')])
        ws_summary.append(["Target Table:", comp_data.get('target_table', 'N/A')])
        ws_summary.append([])
        
        diffs = comp_data.get('differences', {})
        ws_summary.append(["Comparison Results"])
        for cell in ws_summary[ws_summary.max_row]:
            cell.fill = section_fill
            cell.font = section_font
        ws_summary.merge_cells(f'A{ws_summary.max_row}:B{ws_summary.max_row}')
        
        ws_summary.append(["Total Differences:", diffs.get('total_count', 0)])
        ws_summary.append(["Structure Differences:", len(diffs.get('structure', []))])
        ws_summary.append(["Index Differences:", len(diffs.get('indexes', []))])
        ws_summary.append(["Constraint Differences:", len(diffs.get('constraints', []))])
        
        row_count = diffs.get('row_count', {})
        if row_count:
            ws_summary.append(["Row Count - Source:", row_count.get('source', 'N/A')])
            ws_summary.append(["Row Count - Target:", row_count.get('target', 'N/A')])
            ws_summary.append(["Row Count Different:", 'Yes' if row_count.get('different') else 'No'])
        
        # Auto-adjust column widths
        for column in ws_summary.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws_summary.column_dimensions[column_letter].width = min(max(max_length + 2, 15), 50)
        
        # Structure Differences Sheet
        if diffs.get('structure'):
            ws_struct = wb.create_sheet("Structure Differences")
            ws_struct.append(["Structure Differences"])
            ws_struct.merge_cells('A1:E1')
            ws_struct['A1'].font = Font(bold=True, size=14, color="4472C4")
            ws_struct['A1'].alignment = Alignment(horizontal="center")
            ws_struct.append([])
            
            ws_struct.append(["Column Name", "Difference Type", "Source", "Target", "Status"])
            for cell in ws_struct[ws_struct.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for diff in diffs['structure']:
                row_data = [
                    diff.get('column_name'),
                    diff.get('diff_type'),
                    diff.get('source_value', 'N/A'),
                    diff.get('target_value', 'N/A'),
                    diff.get('type', 'modified').upper()
                ]
                ws_struct.append(row_data)
                
                # Apply color coding
                row_num = ws_struct.max_row
                fill_color = modified_fill
                if diff.get('type') == 'added':
                    fill_color = added_fill
                elif diff.get('type') == 'removed':
                    fill_color = removed_fill
                
                for col in range(1, 6):
                    ws_struct.cell(row_num, col).fill = fill_color
                    ws_struct.cell(row_num, col).border = border
            
            # Auto-adjust column widths
            for column in ws_struct.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_struct.column_dimensions[column_letter].width = min(max(max_length + 2, 15), 50)
        
        # Row Count Sheet
        if row_count and row_count.get('different'):
            ws_rows = wb.create_sheet("Row Count Difference")
            ws_rows.append(["Row Count Difference"])
            ws_rows.merge_cells('A1:C1')
            ws_rows['A1'].font = Font(bold=True, size=14, color="4472C4")
            ws_rows['A1'].alignment = Alignment(horizontal="center")
            ws_rows.append([])
            
            ws_rows.append(["Metric", "Value"])
            for cell in ws_rows[ws_rows.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            ws_rows.append(["Source Rows", row_count.get('source')])
            ws_rows.append(["Target Rows", row_count.get('target')])
            ws_rows.append(["Difference", row_count.get('difference')])
            
            for column in ws_rows.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_rows.column_dimensions[column_letter].width = min(max(max_length + 2, 20), 50)
        
        # Index Differences Sheet
        if diffs.get('indexes'):
            ws_idx = wb.create_sheet("Index Differences")
            ws_idx.append(["Index Differences"])
            ws_idx.merge_cells('A1:D1')
            ws_idx['A1'].font = Font(bold=True, size=14, color="4472C4")
            ws_idx['A1'].alignment = Alignment(horizontal="center")
            ws_idx.append([])
            
            ws_idx.append(["Index Name", "Type", "Status"])
            for cell in ws_idx[ws_idx.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for diff in diffs['indexes']:
                row_data = [
                    diff.get('index_name'),
                    diff.get('diff_type'),
                    diff.get('status', '').replace('_', ' ').upper()
                ]
                ws_idx.append(row_data)
                
                # Apply color coding
                row_num = ws_idx.max_row
                fill_color = modified_fill
                if 'missing_in_source' in diff.get('status', ''):
                    fill_color = added_fill
                elif 'missing_in_target' in diff.get('status', ''):
                    fill_color = removed_fill
                
                for col in range(1, 4):
                    ws_idx.cell(row_num, col).fill = fill_color
                    ws_idx.cell(row_num, col).border = border
            
            for column in ws_idx.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_idx.column_dimensions[column_letter].width = min(max(max_length + 2, 20), 50)
        
        # Constraint Differences Sheet
        if diffs.get('constraints'):
            ws_const = wb.create_sheet("Constraint Differences")
            ws_const.append(["Constraint Differences"])
            ws_const.merge_cells('A1:F1')
            ws_const['A1'].font = Font(bold=True, size=14, color="4472C4")
            ws_const['A1'].alignment = Alignment(horizontal="center")
            ws_const.append([])
            
            ws_const.append(["Constraint Name", "Type", "Source Value", "Target Value", "Status"])
            for cell in ws_const[ws_const.max_row]:
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
            
            for diff in diffs['constraints']:
                row_data = [
                    diff.get('constraint_name'),
                    diff.get('constraint_type'),
                    diff.get('source_value', 'N/A'),
                    diff.get('target_value', 'N/A'),
                    diff.get('status', '').replace('_', ' ').upper()
                ]
                ws_const.append(row_data)
                
                # Apply color coding
                row_num = ws_const.max_row
                fill_color = modified_fill
                if 'missing_in_source' in diff.get('status', ''):
                    fill_color = added_fill
                elif 'missing_in_target' in diff.get('status', ''):
                    fill_color = removed_fill
                elif 'modified' in diff.get('status', ''):
                    fill_color = modified_fill
                
                for col in range(1, 6):
                    ws_const.cell(row_num, col).fill = fill_color
                    ws_const.cell(row_num, col).border = border
            
            for column in ws_const.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws_const.column_dimensions[column_letter].width = min(max(max_length + 2, 20), 50)
        
        # Save to BytesIO
        output_file = io.BytesIO()
        wb.save(output_file)
        output_file.seek(0)
        
        # Generate filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"table_comparison_{timestamp}.xlsx"
        
        logger.info(f"Table comparison exported successfully: {filename}")
        
        return send_file(
            output_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as error:
        logger.error(f"Error exporting comparison: {str(error)}")
        return jsonify({'success': False, 'error': str(error)}), 500

@app.route('/api/compare-query-dual', methods=['POST'])
def compare_query_dual():
    """Compare SQL query results from two different databases"""
    try:
        data = request.get_json()
        source_session = data.get('source_session')
        target_session = data.get('target_session')
        source_query = data.get('source_query', '').strip()
        target_query = data.get('target_query', '').strip()
        
        if source_session not in _dual_connections:
            return jsonify({'success': False, 'error': 'Source connection not found'}), 400
        
        if target_session not in _dual_connections:
            return jsonify({'success': False, 'error': 'Target connection not found'}), 400
        
        if not source_query or not target_query:
            return jsonify({'success': False, 'error': 'Both queries are required'}), 400
        
        source_conn = _dual_connections[source_session]['connection']
        target_conn = _dual_connections[target_session]['connection']
        
        source_cursor = source_conn.cursor()
        target_cursor = target_conn.cursor()
        
        try:
            # Execute source query
            source_cursor.execute(source_query)
            source_rows = source_cursor.fetchall()
            source_cols = [desc[0] for desc in source_cursor.description] if source_cursor.description else []
            
            # Execute target query
            target_cursor.execute(target_query)
            target_rows = target_cursor.fetchall()
            target_cols = [desc[0] for desc in target_cursor.description] if target_cursor.description else []
            
            # STRICT VALIDATION: Check column structure first
            column_diffs = []
            
            # Check if column counts differ
            if len(source_cols) != len(target_cols):
                column_diffs.append({
                    'type': 'COLUMN_COUNT_MISMATCH',
                    'message': f'Column count mismatch: Source has {len(source_cols)} columns, Target has {len(target_cols)} columns',
                    'source_columns': ', '.join(source_cols),
                    'target_columns': ', '.join(target_cols)
                })
            
            # Check if column names differ
            source_cols_set = set(source_cols)
            target_cols_set = set(target_cols)
            
            missing_in_target = source_cols_set - target_cols_set
            missing_in_source = target_cols_set - source_cols_set
            
            for col in missing_in_target:
                column_diffs.append({
                    'type': 'COLUMN_MISSING_IN_TARGET',
                    'column_name': col,
                    'message': f'Column "{col}" exists in source but not in target'
                })
            
            for col in missing_in_source:
                column_diffs.append({
                    'type': 'COLUMN_MISSING_IN_SOURCE',
                    'column_name': col,
                    'message': f'Column "{col}" exists in target but not in source'
                })
            
            # Check if column order differs (even if same columns)
            if source_cols_set == target_cols_set and source_cols != target_cols:
                column_diffs.append({
                    'type': 'COLUMN_ORDER_DIFF',
                    'message': 'Columns exist in both but in different order',
                    'source_order': ', '.join(source_cols),
                    'target_order': ', '.join(target_cols)
                })
            
            # Convert to dictionaries for data comparison
            source_data = [dict(zip(source_cols, row)) for row in source_rows]
            target_data = [dict(zip(target_cols, row)) for row in target_rows]
            
            # STRICT VALIDATION: Compare row counts
            row_count_diff = len(source_data) != len(target_data)
            
            # Compare data rows
            matching_rows = 0
            row_differences = []
            rows_only_in_source = []
            rows_only_in_target = []
            
            # Get common columns for data comparison
            common_cols = list(source_cols_set & target_cols_set)
            
            # Compare rows that exist in both
            max_rows = max(len(source_data), len(target_data))
            
            for i in range(max_rows):
                if i < len(source_data) and i < len(target_data):
                    # Both have this row - compare values
                    s_row = source_data[i]
                    t_row = target_data[i]
                    row_match = True
                    
                    # Compare all columns (not just common ones)
                    all_cols = source_cols_set | target_cols_set
                    
                    for col in all_cols:
                        s_val = s_row.get(col, None)
                        t_val = t_row.get(col, None)
                        
                        # Handle None values properly
                        if s_val is None and t_val is None:
                            continue
                        
                        # Convert to string for comparison (handles different types)
                        s_str = str(s_val) if s_val is not None else 'NULL'
                        t_str = str(t_val) if t_val is not None else 'NULL'
                        
                        if col not in s_row:
                            # Column doesn't exist in source
                            row_match = False
                            row_differences.append({
                                'row_number': i + 1,
                                'column_name': col,
                                'source_value': 'COLUMN_NOT_IN_SOURCE',
                                'target_value': t_str,
                                'diff_type': 'column_missing'
                            })
                        elif col not in t_row:
                            # Column doesn't exist in target
                            row_match = False
                            row_differences.append({
                                'row_number': i + 1,
                                'column_name': col,
                                'source_value': s_str,
                                'target_value': 'COLUMN_NOT_IN_TARGET',
                                'diff_type': 'column_missing'
                            })
                        elif s_str != t_str:
                            # Values differ
                            row_match = False
                            row_differences.append({
                                'row_number': i + 1,
                                'column_name': col,
                                'source_value': s_str,
                                'target_value': t_str,
                                'diff_type': 'value_diff'
                            })
                    
                    if row_match:
                        matching_rows += 1
                
                elif i < len(source_data):
                    # Row exists only in source
                    rows_only_in_source.append(i + 1)
                else:
                    # Row exists only in target
                    rows_only_in_target.append(i + 1)
            
            # Calculate total differences
            total_diffs = (
                len(column_diffs) +
                len(row_differences) +
                len(rows_only_in_source) +
                len(rows_only_in_target)
            )
            
            # Add row count difference if exists
            if row_count_diff:
                total_diffs += 1
            
            return jsonify({
                'success': True,
                'summary': {
                    'source_rows': len(source_data),
                    'target_rows': len(target_data),
                    'source_columns': len(source_cols),
                    'target_columns': len(target_cols),
                    'matching_rows': matching_rows,
                    'total_differences': total_diffs,
                    'has_column_differences': len(column_diffs) > 0,
                    'has_row_count_difference': row_count_diff
                },
                'differences': {
                    'column_structure': column_diffs,
                    'row_differences': row_differences,
                    'rows_only_in_source': rows_only_in_source,
                    'rows_only_in_target': rows_only_in_target
                }
            })
            
        finally:
            source_cursor.close()
            target_cursor.close()
            
    except Exception as e:
        logger.error(f"Error in compare-query-dual: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500


def perform_aggressive_optimization(query, db_type, options):
    """
    Perform aggressive query rewriting for actual optimization
    Returns runnable, optimized SQL
    """
    import re
    
    optimized = query
    changes_made = []
    
    # 1. CONVERT CORRELATED SUBQUERIES TO WINDOW FUNCTIONS
    # Pattern: WHERE column > (SELECT AVG(column) FROM same_table WHERE date_match)
    # This is your exact use case!
    
    # Detect pattern: WHERE amount > (SELECT AVG(amount) FROM sales s2 WHERE DATE_FORMAT...)
    correlated_pattern = r'SELECT\s+(.*?)\s+FROM\s+(\w+)\s+(\w+)\s+WHERE\s+(\w+)\s*>\s*\(\s*SELECT\s+AVG\((\w+)\)\s+FROM\s+(\w+)\s+(\w+)\s+WHERE\s+(.*?)\)'
    
    match = re.search(correlated_pattern, query, re.IGNORECASE | re.DOTALL)
    if match and options.get('subqueries', True):
        select_cols = match.group(1).strip()
        table1 = match.group(2)
        alias1 = match.group(3)
        compare_col = match.group(4)
        avg_col = match.group(5)
        table2 = match.group(6)
        alias2 = match.group(7)
        where_condition = match.group(8).strip()
        
        # Check if it's same table
        if table1.upper() == table2.upper():
            # Extract the partitioning condition
            partition_col = None
            
            # Look for DATE_FORMAT or TRUNC patterns
            date_format_match = re.search(r'DATE_FORMAT\((\w+)\.(\w+),\s*[\'"]([^\'"]+)[\'"]\)\s*=\s*DATE_FORMAT\((\w+)\.(\w+),\s*[\'"]([^\'"]+)[\'"]\)', where_condition, re.IGNORECASE)
            
            if date_format_match:
                partition_col = date_format_match.group(2)  # sale_date
                date_format_str = date_format_match.group(3)  # '%Y-%m'
                
                # Rewrite using window function
                if db_type in ['mysql', 'databricks', 'snowflake', 'postgresql']:
                    # MySQL 8+, Databricks, Snowflake, PostgreSQL support window functions
                    optimized = f"""SELECT {select_cols}
FROM (
    SELECT {select_cols},
           AVG({avg_col}) OVER (PARTITION BY DATE_FORMAT({partition_col}, '{date_format_str}')) as month_avg
    FROM {table1}
) {alias1}
WHERE {compare_col} > month_avg"""
                    
                    changes_made.append({
                        'type': 'CORRELATED_SUBQUERY_TO_WINDOW',
                        'description': 'Converted correlated subquery to window function',
                        'improvement': 'Eliminates row-by-row execution, runs in single pass'
                    })
                
                elif db_type == 'oracle':
                    # Oracle version with TRUNC
                    optimized = f"""SELECT {select_cols}
FROM (
    SELECT {select_cols},
           AVG({avg_col}) OVER (PARTITION BY TRUNC({partition_col}, 'MM')) as month_avg
    FROM {table1}
) {alias1}
WHERE {compare_col} > month_avg"""
                    
                    changes_made.append({
                        'type': 'CORRELATED_SUBQUERY_TO_WINDOW',
                        'description': 'Converted correlated subquery to analytic function',
                        'improvement': 'Eliminates row-by-row execution, uses Oracle analytics'
                    })
    
    # 2. CONVERT IN (SELECT...) TO EXISTS
    if options.get('subqueries', True):
        # Pattern: WHERE column IN (SELECT column FROM table WHERE condition)
        in_subquery_pattern = r'WHERE\s+([\w.]+)\s+IN\s*\(\s*SELECT\s+([\w.]+)\s+FROM\s+([\w\s]+?)(?:WHERE\s+(.*?))?\s*\)'
        
        def replace_in_with_exists(match):
            outer_col = match.group(1)
            inner_col = match.group(2)
            inner_table = match.group(3).strip()
            inner_where = match.group(4) if match.group(4) else None
            
            # Generate table alias if not present
            if ' ' not in inner_table:
                inner_table = f"{inner_table} sub"
            
            table_parts = inner_table.split()
            table_alias = table_parts[-1] if len(table_parts) > 1 else 'sub'
            
            exists_clause = f"WHERE EXISTS (SELECT 1 FROM {inner_table} WHERE {table_alias}.{inner_col.split('.')[-1]} = {outer_col}"
            if inner_where:
                exists_clause += f" AND {inner_where}"
            exists_clause += ")"
            
            return exists_clause
        
        if re.search(in_subquery_pattern, optimized, re.IGNORECASE | re.DOTALL):
            optimized = re.sub(in_subquery_pattern, replace_in_with_exists, optimized, flags=re.IGNORECASE | re.DOTALL)
            changes_made.append({
                'type': 'IN_TO_EXISTS',
                'description': 'Converted IN subquery to EXISTS',
                'improvement': 'EXISTS can short-circuit, often faster than IN'
            })
    
    # 3. REMOVE SELECT * and replace with actual column suggestion
    if options.get('bestPractices', True):
        # For demonstration, we'll just add a clear comment, not remove it
        # In production, you'd need to actually query metadata to get column names
        select_star_pattern = r'SELECT\s+\*\s+FROM'
        if re.search(select_star_pattern, optimized, re.IGNORECASE):
            # Don't actually change * to columns (we don't know the schema)
            # But we can suggest it clearly
            pass  # Keeping suggestions in the suggestion section
    
    return optimized, changes_made

def call_ai_model_for_optimization(query: str, db_type: str) -> Dict[str, Any]:
    """
    Call live AI model API (OpenAI GPT or Google Gemini) for SQL query optimization
    
    Args:
        query: The SQL query to optimize
        db_type: Database type (oracle, mysql, postgresql, etc.)
    
    Returns:
        Dict containing optimized query and explanation
    """
    if not AI_MODEL_AVAILABLE:
        return {
            'success': False,
            'error': 'AI Model not configured. Please set AI_API_KEY environment variable.',
            'optimized_query': query,
            'explanation': 'AI optimization unavailable - using rule-based optimization only'
        }
    
    try:
        # Create the prompt for the AI model
        prompt = f"""You are an expert SQL query optimization assistant. Analyze and optimize the following SQL query for {db_type.upper()} database.

Original Query:
```sql
{query}
```

Please provide:
1. An optimized version of the query
2. Detailed explanation of all optimizations made
3. Performance improvement estimation
4. Index recommendations if applicable
5. Best practices applied

Format your response as JSON with these fields:
{{
    "optimized_query": "the optimized SQL query",
    "explanation": "detailed explanation of changes",
    "improvements": ["list of improvements made"],
    "performance_gain": "estimated performance improvement percentage",
    "index_recommendations": ["list of index recommendations"],
    "best_practices": ["list of best practices applied"]
}}

Important: Ensure the optimized query is syntactically correct and runnable."""

        if AI_MODEL_PROVIDER == 'openai':
            # Check rate limit BEFORE making the API call
            rate_limit_check = check_rate_limit('openai')
            
            if not rate_limit_check['allowed']:
                wait_time = rate_limit_check['wait_time']
                error_msg = f"Rate limit exceeded. Please wait {wait_time:.1f} seconds before trying again. Free tier: 3 requests/min."
                logger.warning(error_msg)
                return {
                    'success': False,
                    'error': error_msg,
                    'optimized_query': query,
                    'explanation': 'Rate limit reached - using rule-based optimization',
                    'wait_time': wait_time
                }
            
            # Call OpenAI API
            headers = {
                'Content-Type': 'application/json',
                'Authorization': f'Bearer {AI_API_KEY}'
            }
            
            payload = {
                'model': OPENAI_MODEL,
                'messages': [
                    {
                        'role': 'system',
                        'content': 'You are an expert SQL database optimization assistant. You provide optimized SQL queries and explain your optimizations clearly.'
                    },
                    {
                        'role': 'user',
                        'content': prompt
                    }
                ],
                'temperature': 0.3,  # Lower temperature for more consistent, focused responses
                'max_tokens': 2000
            }
            
            logger.info(f"Calling OpenAI API for query optimization... (Remaining requests: {get_rate_limit_status('openai')['remaining_requests']})")

            # Create a session with retry logic and SSL bypass
            session = requests.Session()
            session.verify = False  # Disable SSL verification

            # Disable SSL warnings for this session
            import warnings
            from urllib3.exceptions import InsecureRequestWarning
            warnings.filterwarnings('ignore', category=InsecureRequestWarning)

            try:
                response = session.post(
                    OPENAI_API_URL,
                    headers=headers,
                    json=payload,
                    timeout=30
                )
                
                # Handle different error codes with helpful messages
                if response.status_code == 429:
                    error_msg = "Rate limit exceeded. Free tier: 3 requests/min. Try again in a minute or upgrade your plan."
                    logger.warning(error_msg)
                    return {
                        'success': False,
                        'error': error_msg,
                        'optimized_query': query,
                        'explanation': 'Rate limit reached - using rule-based optimization'
                    }
                elif response.status_code == 401:
                    error_msg = "Invalid API key. Please check your OpenAI API key."
                    logger.error(error_msg)
                    return {
                        'success': False,
                        'error': error_msg,
                        'optimized_query': query,
                        'explanation': 'Authentication failed - using rule-based optimization'
                    }
                elif response.status_code == 403:
                    error_msg = "Access forbidden. Your API key may not have permission for this model."
                    logger.error(error_msg)
                    return {
                        'success': False,
                        'error': error_msg,
                        'optimized_query': query,
                        'explanation': 'Permission denied - using rule-based optimization'
                    }
                
                response.raise_for_status()
                
                # Record successful API call for rate limiting
                record_api_call('openai')
                logger.info(f"API call successful. Rate limit status: {get_rate_limit_status('openai')}")

                result = response.json()
                ai_response = result['choices'][0]['message']['content']
            finally:
                session.close()

        elif AI_MODEL_PROVIDER == 'gemini':
            # Call Google Gemini API
            logger.info(f"Calling Google Gemini API for query optimization...")
            
            headers = {
                'Content-Type': 'application/json'
            }
            
            payload = {
                'contents': [{
                    'parts': [{
                        'text': prompt
                    }]
                }],
                'generationConfig': {
                    'temperature': 0.3,
                    'maxOutputTokens': 2000
                }
            }
            
            api_url = f"{GEMINI_API_URL}?key={AI_API_KEY}"
            
            # Disable SSL verification to avoid certificate errors (use cautiously)
            response = requests.post(api_url, headers=headers, json=payload, timeout=30, verify=False)
            response.raise_for_status()
            
            result = response.json()
            ai_response = result['candidates'][0]['content']['parts'][0]['text']
        
        else:
            return {
                'success': False,
                'error': f'Unsupported AI provider: {AI_MODEL_PROVIDER}',
                'optimized_query': query,
                'explanation': 'Invalid AI provider configuration'
            }
        
        # Parse AI response
        # Try to extract JSON from the response
        try:
            # Look for JSON block in markdown code fence or plain JSON
            import re
            json_match = re.search(r'```(?:json)?\s*(\{.*\})\s*```', ai_response, re.DOTALL)
            if json_match:
                ai_data = json.loads(json_match.group(1))
            else:
                # Try to parse the entire response as JSON
                ai_data = json.loads(ai_response)
        except json.JSONDecodeError:
            # If JSON parsing fails, use the response as-is with basic structure
            logger.warning("AI response is not valid JSON, using fallback parsing")
            ai_data = {
                'optimized_query': query,  # Keep original if parsing fails
                'explanation': ai_response,
                'improvements': ['AI optimization applied'],
                'performance_gain': 'See explanation',
                'index_recommendations': [],
                'best_practices': []
            }
        
        return {
            'success': True,
            'optimized_query': ai_data.get('optimized_query', query),
            'explanation': ai_data.get('explanation', ''),
            'improvements': ai_data.get('improvements', []),
            'performance_gain': ai_data.get('performance_gain', 'Unknown'),
            'index_recommendations': ai_data.get('index_recommendations', []),
            'best_practices': ai_data.get('best_practices', []),
            'ai_provider': AI_MODEL_PROVIDER
        }
        
    except requests.exceptions.Timeout:
        logger.error("AI API request timed out")
        return {
            'success': False,
            'error': 'AI API request timed out. Please try again.',
            'optimized_query': query,
            'explanation': 'Request timeout'
        }
    except requests.exceptions.RequestException as e:
        logger.error(f"AI API request failed: {str(e)}")
        return {
            'success': False,
            'error': f'AI API request failed: {str(e)}',
            'optimized_query': query,
            'explanation': 'API request failed'
        }
    except Exception as e:
        logger.error(f"AI optimization error: {str(e)}")
        return {
            'success': False,
            'error': f'AI optimization error: {str(e)}',
            'optimized_query': query,
            'explanation': 'Unexpected error during AI optimization'
        }


@app.route('/api/optimize-query', methods=['POST'])
def optimize_query():
    """
    AI-Powered SQL Query Optimizer
    Analyzes SQL queries and provides optimization suggestions
    Uses live AI model (OpenAI GPT or Google Gemini) when available
    """
    try:
        data = request.get_json()
        original_query = data.get('query', '').strip()
        db_type = data.get('db_type', 'oracle').lower()
        options = data.get('options', {})
        use_ai = data.get('use_ai', True)  # Enable AI by default if available
        
        if not original_query:
            return jsonify({'success': False, 'error': 'No query provided'}), 400
        
        # Initialize optimization results
        suggestions = []
        improvements = []
        index_recommendations = []
        optimizations_count = 0
        ai_used = False
        
        # Convert query to uppercase for analysis
        query_upper = original_query.upper()
        
        # STEP 0: TRY AI MODEL OPTIMIZATION FIRST (if enabled and available)
        if use_ai and AI_MODEL_AVAILABLE:
            logger.info("Using live AI model for query optimization...")
            ai_result = call_ai_model_for_optimization(original_query, db_type)
            
            if ai_result.get('success'):
                ai_used = True
                optimized_query = ai_result.get('optimized_query', original_query)
                
                # Add AI-generated improvements
                if ai_result.get('improvements'):
                    for improvement in ai_result['improvements']:
                        improvements.append({
                            'title': '🤖 AI Optimization',
                            'description': improvement,
                            'impact': 'AI-recommended enhancement'
                        })
                        optimizations_count += 1
                
                # Add AI explanation as high-priority suggestion
                if ai_result.get('explanation'):
                    suggestions.append({
                        'priority': 'high',
                        'title': '🤖 AI Analysis',
                        'description': ai_result['explanation'],
                        'example': 'See optimized query above'
                    })
                
                # Add AI index recommendations
                if ai_result.get('index_recommendations'):
                    for idx_rec in ai_result['index_recommendations']:
                        index_recommendations.append({
                            'table': 'AI-suggested',
                            'columns': 'See recommendation',
                            'type': 'INDEX',
                            'reason': idx_rec,
                            'ddl': idx_rec
                        })
                
                # Add AI best practices
                if ai_result.get('best_practices'):
                    for practice in ai_result['best_practices']:
                        suggestions.append({
                            'priority': 'info',
                            'title': '✅ Best Practice',
                            'description': practice,
                            'example': 'Applied in optimized query'
                        })

                # Calculate AI-based performance gain
                performance_gain = ai_result.get('performance_gain', 'Unknown')

                logger.info(f"AI optimization completed successfully using {ai_result.get('ai_provider', 'unknown')}")
            else:
                # AI failed, fall back to rule-based
                logger.warning(f"AI optimization failed: {ai_result.get('error', 'Unknown error')}, using rule-based optimization")
                optimized_query = original_query
                suggestions.append({
                    'priority': 'info',
                    'title': '⚠️ AI Optimization Unavailable',
                    'description': ai_result.get('error', 'AI model is not available'),
                    'example': 'Using rule-based optimization instead'
                })
        else:
            # AI not available or not requested, use rule-based optimization
            optimized_query = original_query
            if not AI_MODEL_AVAILABLE:
                suggestions.append({
                    'priority': 'info',
                    'title': 'ℹ️ AI Model Not Configured',
                    'description': 'Set AI_API_KEY environment variable to enable AI-powered optimization. Using rule-based optimization instead.',
                    'example': 'export AI_API_KEY="your-api-key-here"'
                })

        # RULE-BASED OPTIMIZATION (Always run if AI is not used or as enhancement)
        # STEP 1: AGGRESSIVE OPTIMIZATION - Actually rewrite the query
        # Only run rule-based if AI was not successful
        if not ai_used:
            optimized_query, aggressive_changes = perform_aggressive_optimization(original_query, db_type, options)
        else:
            # AI was used, skip aggressive optimization to avoid conflicts
            aggressive_changes = []

        # Add aggressive optimization results
        for change in aggressive_changes:
            optimizations_count += 1
            improvements.append({
                'title': change['description'],
                'description': change['description'],
                'impact': change['improvement']
            })
            suggestions.append({
                'priority': 'high',
                'title': change['type'].replace('_', ' ').title(),
                'description': change['improvement'],
                'example': 'See optimized query'
            })
        
        # 1. DETECT AND OPTIMIZE IMPLICIT JOINS (Cartesian Products)
        if options.get('joins', True):
            if 'FROM' in query_upper and ',' in query_upper and 'JOIN' not in query_upper:
                # Detected implicit join (old-style comma join)
                suggestions.append({
                    'priority': 'high',
                    'title': 'Implicit Join Detected',
                    'description': 'Old-style comma joins (FROM table1, table2) should be replaced with explicit JOIN syntax for better readability and optimizer hints.',
                    'example': 'FROM employees e JOIN departments d ON e.dept_id = d.id'
                })
                
                # Convert comma joins to INNER JOIN
                optimized_query = convert_implicit_to_explicit_joins(optimized_query, db_type)
                optimizations_count += 1
                
                improvements.append({
                    'title': 'Converted Implicit Joins to Explicit JOIN Syntax',
                    'description': 'Replaced old-style comma joins with modern INNER JOIN syntax',
                    'impact': 'Better query readability and optimizer control'
                })
        
        # 2. ADD QUERY HINTS BASED ON DATABASE TYPE
        if options.get('hints', True):
            hints_added = add_database_hints(optimized_query, db_type, query_upper)
            if hints_added['modified']:
                optimized_query = hints_added['query']
                optimizations_count += hints_added['count']
                suggestions.extend(hints_added['suggestions'])
                improvements.extend(hints_added['improvements'])
        
        # 3. OPTIMIZE SELECT * USAGE
        if options.get('bestPractices', True):
            if 'SELECT *' in query_upper or 'SELECT*' in query_upper.replace(' ', ''):
                # Add suggestion only - don't modify query
                suggestions.append({
                    'priority': 'medium',
                    'title': 'Avoid SELECT * in Production Queries',
                    'description': 'SELECT * retrieves all columns, which can be inefficient. Specify only the columns you need.',
                    'example': 'SELECT column1, column2, column3 FROM table_name'
                })
        
        # 4. INDEX RECOMMENDATIONS
        if options.get('indexes', True):
            index_recs = analyze_index_opportunities(original_query, db_type)
            index_recommendations.extend(index_recs)
            if index_recs:
                optimizations_count += len(index_recs)
        
        # 5. SUBQUERY OPTIMIZATION (suggestions only, actual rewrite done in aggressive optimization)
        if options.get('subqueries', True):
            if query_upper.count('SELECT') > 1:  # Has subqueries
                subquery_opts = optimize_subqueries(optimized_query, db_type, query_upper)
                if subquery_opts['suggestions']:
                    suggestions.extend(subquery_opts['suggestions'])
        
        # 6. WHERE CLAUSE OPTIMIZATION (suggestions only)
        if options.get('bestPractices', True):
            where_opts = analyze_where_clause(query_upper, optimized_query)
            suggestions.extend(where_opts['suggestions'])
        
        # 7. ORDER BY / GROUP BY OPTIMIZATION
        if 'ORDER BY' in query_upper or 'GROUP BY' in query_upper:
            sort_opts = optimize_sorting(query_upper, db_type)
            suggestions.extend(sort_opts)
        
        # 8. CALCULATE COMPLEXITY AND ESTIMATED IMPROVEMENT
        complexity = calculate_query_complexity(query_upper)
        estimated_improvement = calculate_estimated_improvement(
            original_query, optimized_query, suggestions, improvements
        )
        
        # Return optimization results
        return jsonify({
            'success': True,
            'original_query': original_query,
            'optimized_query': optimized_query,
            'db_type': db_type,
            'optimizations_count': optimizations_count,
            'estimated_improvement': estimated_improvement if not ai_used else performance_gain,
            'complexity': complexity,
            'suggestions': suggestions,
            'improvements': improvements,
            'index_recommendations': index_recommendations,
            'ai_used': ai_used,
            'ai_provider': AI_MODEL_PROVIDER if ai_used else None,
            'ai_available': AI_MODEL_AVAILABLE
        })
        
    except Exception as e:
        logger.error(f"Query optimization error: {str(e)}")
        return jsonify({
            'success': False,
            'error': f'Optimization error: {str(e)}'
        }), 500


@app.route('/api/export-optimized-query', methods=['POST'])
def export_optimized_query():
    """
    Export optimized SQL query to a .txt file
    Includes query, metadata, and optimization details
    """
    try:
        data = request.get_json()
        original_query = data.get('original_query', '')
        optimized_query = data.get('optimized_query', '')
        db_type = data.get('db_type', 'unknown')
        improvements = data.get('improvements', [])
        suggestions = data.get('suggestions', [])
        index_recommendations = data.get('index_recommendations', [])
        ai_used = data.get('ai_used', False)
        ai_provider = data.get('ai_provider', 'N/A')
        estimated_improvement = data.get('estimated_improvement', 'Unknown')
        
        if not optimized_query:
            return jsonify({'success': False, 'error': 'No optimized query provided'}), 400
        
        # Create formatted text content
        content = f"""{'='*80}
SQL QUERY OPTIMIZATION REPORT
{'='*80}

Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
Database Type: {db_type.upper()}
Optimization Method: {'🤖 AI-Powered (' + ai_provider.upper() + ')' if ai_used else '📋 Rule-Based'}
Estimated Performance Improvement: {estimated_improvement}

{'='*80}
ORIGINAL QUERY
{'='*80}

{original_query}

{'='*80}
OPTIMIZED QUERY
{'='*80}

{optimized_query}

{'='*80}
OPTIMIZATION IMPROVEMENTS
{'='*80}

"""
        # Add improvements
        if improvements:
            for idx, improvement in enumerate(improvements, 1):
                content += f"{idx}. {improvement.get('title', 'Optimization')}\n"
                content += f"   Description: {improvement.get('description', 'N/A')}\n"
                content += f"   Impact: {improvement.get('impact', 'N/A')}\n\n"
        else:
            content += "No specific improvements recorded.\n\n"
        
        # Add suggestions
        if suggestions:
            content += f"{'='*80}\n"
            content += "OPTIMIZATION SUGGESTIONS\n"
            content += f"{'='*80}\n\n"
            
            for idx, suggestion in enumerate(suggestions, 1):
                priority = suggestion.get('priority', 'info').upper()
                content += f"{idx}. [{priority}] {suggestion.get('title', 'Suggestion')}\n"
                content += f"   {suggestion.get('description', 'N/A')}\n"
                if suggestion.get('example'):
                    content += f"   Example: {suggestion['example']}\n"
                content += "\n"
        
        # Add index recommendations
        if index_recommendations:
            content += f"{'='*80}\n"
            content += "INDEX RECOMMENDATIONS\n"
            content += f"{'='*80}\n\n"
            
            for idx, rec in enumerate(index_recommendations, 1):
                content += f"{idx}. Table: {rec.get('table', 'N/A')}\n"
                content += f"   Columns: {rec.get('columns', 'N/A')}\n"
                content += f"   Type: {rec.get('type', 'INDEX')}\n"
                content += f"   Reason: {rec.get('reason', 'N/A')}\n"
                if rec.get('ddl'):
                    content += f"   DDL: {rec['ddl']}\n"
                content += "\n"
        
        content += f"{'='*80}\n"
        content += "END OF REPORT\n"
        content += f"{'='*80}\n"
        
        # Create BytesIO object with the content
        output = io.BytesIO(content.encode('utf-8'))
        output.seek(0)
        
        # Generate filename
        filename = f"optimized_query_{db_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        # Also save to backup directory
        backup_path = os.path.join(BACKUP_DIR, filename)
        with open(backup_path, 'w', encoding='utf-8') as f:
            f.write(content)
        
        logger.info(f"Optimized query exported to: {backup_path}")
        
        return send_file(
            output,
            mimetype='text/plain',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error exporting optimized query: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def convert_implicit_to_explicit_joins(query, db_type):
    """Convert old-style comma joins to explicit JOIN syntax"""
    import re
    
    # Check if there's a comma join pattern
    if ',' not in query or 'FROM' not in query.upper():
        return query
    
    # Pattern to find FROM clause with comma-separated tables
    from_pattern = r'FROM\s+([\w\s,\.]+?)(?:WHERE|ORDER|GROUP|HAVING|LIMIT|$)'
    from_match = re.search(from_pattern, query, re.IGNORECASE | re.DOTALL)
    
    if not from_match:
        return query
    
    from_clause = from_match.group(1).strip()
    
    # Check if it's really a comma join (not a function parameter)
    if ',' not in from_clause:
        return query
    
    # Split tables by comma
    tables = [t.strip() for t in from_clause.split(',')]
    
    if len(tables) < 2:
        return query
    
    # Parse first table (this becomes the base)
    base_table = tables[0]
    
    # Build JOIN clause
    join_clause = base_table
    
    # Extract WHERE clause to find join conditions
    where_pattern = r'WHERE\s+(.+?)(?:ORDER|GROUP|HAVING|LIMIT|;|$)'
    where_match = re.search(where_pattern, query, re.IGNORECASE | re.DOTALL)
    
    where_conditions = []
    join_conditions = []
    
    if where_match:
        where_clause = where_match.group(1).strip()
        # Split by AND
        conditions = re.split(r'\s+AND\s+', where_clause, flags=re.IGNORECASE)
        
        # Extract table aliases
        table_aliases = []
        for table in tables:
            parts = table.split()
            if len(parts) >= 2:
                table_aliases.append(parts[-1])  # Last part is alias
            else:
                table_aliases.append(parts[0])  # Table name itself
        
        # Separate join conditions from filter conditions
        for condition in conditions:
            # Check if condition involves two table aliases (likely a join condition)
            alias_matches = sum(1 for alias in table_aliases if re.search(rf'\b{alias}\b', condition, re.IGNORECASE))
            if alias_matches >= 2:
                join_conditions.append(condition.strip())
            else:
                where_conditions.append(condition.strip())
    
    # Build INNER JOINs
    for i, table in enumerate(tables[1:], 1):
        # Find matching join condition
        table_parts = table.split()
        table_alias = table_parts[-1] if len(table_parts) >= 2 else table_parts[0]
        
        # Try to find a join condition for this table
        matching_condition = None
        for condition in join_conditions:
            if re.search(rf'\b{table_alias}\b', condition, re.IGNORECASE):
                matching_condition = condition
                break
        
        if matching_condition:
            join_clause += f"\n    INNER JOIN {table} ON {matching_condition}"
            join_conditions.remove(matching_condition)
        else:
            # No specific condition found, use a generic one
            join_clause += f"\n    INNER JOIN {table} ON 1=1  -- TODO: Add proper join condition"
    
    # Rebuild the query
    optimized = query
    
    # Replace FROM clause
    optimized = re.sub(from_pattern, f'FROM {join_clause}\nWHERE ' if where_conditions or join_conditions else f'FROM {join_clause}\n', optimized, flags=re.IGNORECASE | re.DOTALL)
    
    # Rebuild WHERE clause with remaining conditions
    if where_match and (where_conditions or join_conditions):
        remaining_conditions = where_conditions + join_conditions
        new_where = ' AND '.join(remaining_conditions)
        
        # Find and replace the WHERE clause
        where_full_pattern = r'WHERE\s+.+?(?=ORDER|GROUP|HAVING|LIMIT|;|$)'
        if new_where:
            optimized = re.sub(where_full_pattern, f'WHERE {new_where}', optimized, flags=re.IGNORECASE | re.DOTALL)
        else:
            # Remove WHERE clause if no conditions remain
            optimized = re.sub(r'\s*WHERE\s+.+?(?=ORDER|GROUP|HAVING|LIMIT|;|$)', '', optimized, flags=re.IGNORECASE | re.DOTALL)
    
    return optimized if optimized != query else query

def add_database_hints(query, db_type, query_upper):
    """Add database-specific optimization hints"""
    import re
    
    result = {
        'modified': False,
        'query': query,
        'count': 0,
        'suggestions': [],
        'improvements': []
    }
    
    if db_type == 'oracle':
        # Oracle-specific hints
        if 'SELECT' in query_upper and '/*+' not in query_upper:
            hints_to_add = []
            
            # Determine which hints to add based on query pattern
            if 'JOIN' in query_upper:
                hints_to_add.append('USE_HASH')
                result['suggestions'].append({
                    'priority': 'medium',
                    'title': 'Oracle Hash Join Hint Added',
                    'description': 'Added /*+ USE_HASH */ hint for better join performance with large tables',
                    'example': 'SELECT /*+ USE_HASH(e d) */ ...'
                })
            
            # Add PARALLEL hint for complex queries
            if query_upper.count('JOIN') > 1 or 'GROUP BY' in query_upper:
                hints_to_add.append('PARALLEL(4)')
                result['suggestions'].append({
                    'priority': 'info',
                    'title': 'Parallel Query Hint Added',
                    'description': 'Added PARALLEL hint for faster query execution on multi-core systems',
                    'example': 'SELECT /*+ PARALLEL(4) */ ...'
                })
            
            # Actually add the hints to the query
            if hints_to_add:
                hint_string = ' '.join(hints_to_add)
                # Find first SELECT and add hint after it
                select_pattern = r'(\bSELECT\b)'
                result['query'] = re.sub(select_pattern, rf'\1 /*+ {hint_string} */', query, count=1, flags=re.IGNORECASE)
                result['modified'] = True
                result['count'] += len(hints_to_add)
                
                result['improvements'].append({
                    'title': f'Added Oracle Hints: {hint_string}',
                    'description': f'Inserted optimizer hints after SELECT clause',
                    'impact': 'Query optimizer will use suggested execution strategy'
                })
            
            if 'INDEX' not in query_upper and 'WHERE' in query_upper:
                result['suggestions'].append({
                    'priority': 'info',
                    'title': 'Index Hint Suggestion',
                    'description': 'If you have indexes, manually add /*+ INDEX(table_name index_name) */ hint',
                    'example': 'SELECT /*+ INDEX(employees emp_idx) */ ...'
                })
            
    elif db_type == 'sqlserver':
        if 'SELECT' in query_upper and 'WITH (NOLOCK)' not in query_upper:
            result['suggestions'].append({
                'priority': 'info',
                'title': 'SQL Server Table Hints',
                'description': 'Consider table hints like WITH (NOLOCK) for read operations (be aware of dirty reads)',
                'example': 'FROM employees WITH (NOLOCK)'
            })
            
            # Add NOLOCK hint to FROM clause (for read operations only)
            # This is aggressive - only do if explicitly requested
            # Commenting out for safety
            # from_pattern = r'FROM\s+(\w+)(\s+\w+)?(?!\s+WITH)'
            # result['query'] = re.sub(from_pattern, r'FROM \1\2 WITH (NOLOCK)', query, flags=re.IGNORECASE)
    
    elif db_type == 'postgresql':
        if 'LIMIT' not in query_upper and 'ORDER BY' in query_upper:
            result['suggestions'].append({
                'priority': 'low',
                'title': 'Consider Adding LIMIT Clause',
                'description': 'For large result sets, add LIMIT to reduce data transfer',
                'example': 'ORDER BY created_date DESC LIMIT 100'
            })
    
    elif db_type == 'databricks':
        # Databricks-specific optimizations
        if 'SELECT' in query_upper:
            # Databricks supports broadcast joins for small tables
            if 'JOIN' in query_upper:
                result['suggestions'].append({
                    'priority': 'info',
                    'title': 'Databricks Broadcast Join',
                    'description': 'For small dimension tables, use broadcast hint: SELECT /*+ BROADCAST(small_table) */',
                    'example': 'SELECT /*+ BROADCAST(dim_table) */ * FROM fact JOIN dim_table'
                })
    
    elif db_type == 'snowflake':
        # Snowflake automatically optimizes, but we can suggest clustering
        if 'SELECT' in query_upper:
            result['suggestions'].append({
                'priority': 'info',
                'title': 'Snowflake Clustering Keys',
                'description': 'Consider defining clustering keys on frequently filtered columns',
                'example': 'ALTER TABLE table_name CLUSTER BY (date_column, id_column);'
            })
    
    return result

def analyze_index_opportunities(query, db_type):
    """Analyze query and recommend indexes"""
    recommendations = []
    query_upper = query.upper()
    
    # Extract table and column names from WHERE clause
    import re
    
    # Find WHERE conditions
    where_pattern = r'WHERE\s+(\w+)\.(\w+)\s*[=<>]'
    matches = re.findall(where_pattern, query, re.IGNORECASE)
    
    for table_alias, column in matches:
        # Try to find actual table name
        table_pattern = rf'FROM\s+(\w+)\s+{table_alias}'
        table_match = re.search(table_pattern, query, re.IGNORECASE)
        
        table_name = table_match.group(1) if table_match else table_alias
        
        recommendations.append({
            'table': table_name,
            'columns': column,
            'type': 'B-TREE' if db_type in ['oracle', 'postgresql', 'mysql'] else 'INDEX',
            'reason': f'Column used in WHERE clause for filtering',
            'ddl': f'CREATE INDEX idx_{table_name}_{column.lower()} ON {table_name}({column})'
        })
    
    # Find JOIN conditions
    join_pattern = r'ON\s+(\w+)\.(\w+)\s*=\s*(\w+)\.(\w+)'
    join_matches = re.findall(join_pattern, query, re.IGNORECASE)
    
    for alias1, col1, alias2, col2 in join_matches:
        recommendations.append({
            'table': alias1,
            'columns': col1,
            'type': 'B-TREE',
            'reason': 'Column used in JOIN condition',
            'ddl': f'CREATE INDEX idx_{alias1}_{col1.lower()} ON {alias1}({col1})'
        })
    
    # Remove duplicates
    unique_recs = []
    seen = set()
    for rec in recommendations:
        key = (rec['table'], rec['columns'])
        if key not in seen:
            seen.add(key)
            unique_recs.append(rec)
    
    return unique_recs[:5]  # Limit to top 5 recommendations

def optimize_subqueries(query, db_type, query_upper):
    """Provide suggestions for subquery optimization"""
    import re
    
    result = {
        'suggestions': [],
        'optimized_query': query,
        'modified': False
    }
    
    if 'IN (SELECT' in query_upper or 'IN(SELECT' in query_upper.replace(' ', ''):
        result['suggestions'].append({
            'priority': 'medium',
            'title': 'IN Subquery Detected',
            'description': 'IN with subqueries can be slow. Consider using EXISTS or JOIN instead',
            'example': 'WHERE EXISTS (SELECT 1 FROM table WHERE table.id = outer.id)'
        })
    
    if 'NOT IN (SELECT' in query_upper or 'NOT IN(SELECT' in query_upper.replace(' ', ''):
        result['suggestions'].append({
            'priority': 'high',
            'title': 'NOT IN Subquery Detected',
            'description': 'NOT IN with NULL values can cause issues. Use NOT EXISTS or LEFT JOIN with NULL check',
            'example': 'WHERE NOT EXISTS (SELECT 1 FROM table WHERE table.id = outer.id)'
        })
    
    # Detect correlated subqueries in WHERE clause (already handled by aggressive optimization)
    if re.search(r'WHERE\s+.+?\(\s*SELECT.+?WHERE.+?\)', query_upper, re.DOTALL):
        result['suggestions'].append({
            'priority': 'high',
            'title': 'Correlated Subquery Detected',
            'description': 'Correlated subqueries execute once per row. Window functions are much faster',
            'example': 'SELECT * FROM (SELECT col, AVG(col) OVER (PARTITION BY group_col) as avg FROM table) WHERE col > avg'
        })
    
    return result

def analyze_where_clause(query_upper, original_query=None):
    """Analyze WHERE clause for optimization opportunities"""
    import re
    
    suggestions = []
    
    if 'OR' in query_upper and 'WHERE' in query_upper:
        suggestions.append({
            'priority': 'medium',
            'title': 'Multiple OR Conditions',
            'description': 'Multiple OR conditions can prevent index usage. Consider using IN or UNION',
            'example': 'WHERE status IN (\'active\', \'pending\') -- instead of: status = \'active\' OR status = \'pending\''
        })
    
    if 'LIKE \'%' in query_upper or 'LIKE "%' in query_upper:
        suggestions.append({
            'priority': 'high',
            'title': 'Leading Wildcard in LIKE',
            'description': 'LIKE with leading wildcard (%) cannot use indexes. Consider full-text search',
            'example': 'Use LIKE \'value%\' (trailing wildcard) or implement full-text search'
        })
    
    # Check for function usage on indexed columns
    func_pattern = r'(UPPER|LOWER|SUBSTR|TO_CHAR|TO_DATE|DATE_FORMAT|TRIM)\s*\(\s*\w+\.\w+\s*\)'
    if re.search(func_pattern, query_upper):
        suggestions.append({
            'priority': 'high',
            'title': 'Function on Column in WHERE Clause',
            'description': 'Functions on columns prevent index usage. Create function-based indexes',
            'example': 'CREATE INDEX idx_func ON table (UPPER(column)); -- then use: WHERE UPPER(column) = \'VALUE\''
        })
    
    return {
        'suggestions': suggestions,
        'optimized_query': original_query if original_query else '',
        'modified': False
    }

def optimize_sorting(query_upper, db_type):
    """Optimize ORDER BY and GROUP BY clauses"""
    suggestions = []
    
    if 'ORDER BY' in query_upper and 'LIMIT' not in query_upper and db_type in ['postgresql', 'mysql']:
        suggestions.append({
            'priority': 'low',
            'title': 'Missing LIMIT with ORDER BY',
            'description': 'ORDER BY without LIMIT sorts entire result set. Add LIMIT for better performance',
            'example': 'ORDER BY created_date DESC LIMIT 100'
        })
    
    if 'GROUP BY' in query_upper and 'HAVING' in query_upper:
        suggestions.append({
            'priority': 'info',
            'title': 'HAVING Clause Usage',
            'description': 'Ensure HAVING filters are necessary. Move conditions to WHERE when possible',
            'example': 'Use WHERE for row-level filtering before GROUP BY'
        })
    
    return suggestions

def calculate_query_complexity(query_upper):
    """Calculate query complexity level"""
    score = 0
    
    # Count joins
    score += query_upper.count('JOIN') * 2
    score += query_upper.count('LEFT JOIN') * 3
    score += query_upper.count('OUTER JOIN') * 3
    
    # Count subqueries
    score += query_upper.count('SELECT') - 1  # Minus main SELECT
    
    # Count aggregations
    score += query_upper.count('GROUP BY') * 2
    score += query_upper.count('HAVING') * 2
    
    # Count conditions
    score += query_upper.count('WHERE')
    score += query_upper.count('AND')
    score += query_upper.count('OR')
    
    if score <= 5:
        return 'Simple'
    elif score <= 15:
        return 'Medium'
    elif score <= 30:
        return 'Complex'
    else:
        return 'Very Complex'

def calculate_estimated_improvement(original, optimized, suggestions, improvements):
    """Calculate estimated performance improvement percentage"""
    improvement = 0
    
    # Base improvement from optimizations applied
    improvement += len(improvements) * 10
    
    # Additional improvement from high-priority suggestions
    high_priority = sum(1 for s in suggestions if s['priority'] == 'high')
    improvement += high_priority * 15
    
    # Medium priority suggestions
    medium_priority = sum(1 for s in suggestions if s['priority'] == 'medium')
    improvement += medium_priority * 10
    
    # Query was modified
    if original != optimized:
        improvement += 20
    
    # Cap at 85%
    return min(improvement, 85)

@app.route('/api/info')
def api_info():
    """API information endpoint"""
    return jsonify({
        'message': 'CrossDB Analyzer API',
        'name': 'CrossDB Analyzer - Cross-Platform Database Analysis & Comparison Tool',
        'developed_by': 'Amdocs Data Team',
        'version': '2.0 - Multi-Database Edition',
        'supported_databases': ['Oracle', 'Databricks', 'Snowflake'],
        'endpoints': {
            'health': '/api/health',
            'login': '/api/login (POST)',
            'logout': '/api/logout (POST)',
            'check_auth': '/api/check-auth (GET)',
            'analyze': '/api/analyze (POST)',
            'export': '/api/export (POST)',
            'compare': '/api/compare (POST)'
        },
        'usage': {
            'single_schema': {
                'tables': ['TABLE1', 'TABLE2'],
                'owner': 'SCHEMA_NAME'
            },
            'multiple_schemas': {
                'tables': ['SCHEMA1.TABLE1', 'SCHEMA2.TABLE2', 'TABLE3'],
                'owner': 'DEFAULT_SCHEMA'
            }
        }
    })

@app.route('/api/rate-limit-status', methods=['GET'])
def rate_limit_status():
    """Return current rate limit status for AI API"""
    try:
        if not AI_MODEL_AVAILABLE:
            return jsonify({
                'success': True,
                'ai_available': False,
                'message': 'AI not configured'
            })
        
        status = get_rate_limit_status(AI_MODEL_PROVIDER)
        
        return jsonify({
            'success': True,
            'ai_available': True,
            'provider': AI_MODEL_PROVIDER,
            'status': status,
            'message': f"{status['remaining_requests']}/{status['max_requests']} requests remaining"
        })
    except Exception as e:
        logger.error(f"Error getting rate limit status: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/analyze', methods=['POST'])
def analyze_tables():
    """
    Main endpoint to analyze tables
    Supports multiple formats:
    1. {"tables": ["TABLE1", "TABLE2"], "owner": "SCHEMA1"}
    2. {"tables": ["SCHEMA1.TABLE1", "SCHEMA2.TABLE2"]}
    3. {"tables": "SCHEMA1.TABLE1, SCHEMA2.TABLE2, TABLE3", "owner": "DEFAULT_SCHEMA"}
    """
    try:
        data = request.get_json()
        table_inputs = data.get('tables', [])
        default_owner = data.get('owner', None)
        
        if not table_inputs:
            return jsonify({'error': 'No table names provided'}), 400
        
        # Handle comma-separated string input
        if isinstance(table_inputs, str):
            table_inputs = [t.strip() for t in table_inputs.split(',')]
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        results = []
        
        for table_input in table_inputs:
            table_input = table_input.strip()
            
            # Parse catalog/database, schema, and table name
            catalog_or_db, schema, table_name = parse_table_input(table_input)
            
            # Use default owner if no schema specified
            if not schema and default_owner:
                schema = default_owner.upper()
            
            # Build full name for logging
            db_type = get_db_type()
            full_name = build_full_table_name(table_name, schema, catalog_or_db, db_type)
            logger.info(f"Analyzing table: {full_name} (DB: {db_type})")
            
            table_info = {
                'table_name': table_name,
                'schema': schema,
                'catalog_or_db': catalog_or_db,
                'count': get_table_count(cursor, table_name, schema, catalog_or_db),
                'structure': get_table_structure(cursor, table_name, schema, catalog_or_db),
                'indexes': get_table_indexes(cursor, table_name, schema, catalog_or_db),
                'partitions': get_table_partitions(cursor, table_name, schema, catalog_or_db),
                'primary_key': get_primary_key(cursor, table_name, schema, catalog_or_db),
                'foreign_keys': get_foreign_keys(cursor, table_name, schema, catalog_or_db),
                'last_analyzed': get_last_analyzed(cursor, table_name, schema, catalog_or_db),
                'grants': get_table_grants(cursor, table_name, schema, catalog_or_db)
            }
            
            results.append(table_info)
        
        cursor.close()
        
        # Don't close cached connections (Azure AD/SSO) to prevent re-authentication
        session_id = session.get('session_id')
        if not (session_id and session_id in _connection_cache):
            connection.close()
        
        return jsonify({
            'success': True,
            'data': results,
            'count': len(results)
        })
        
    except Exception as e:
        logger.error(f"Error in analyze_tables: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/export', methods=['POST'])
def export_to_excel():
    """Export analysis results to Excel and save to backup directory"""
    try:
        data = request.get_json()
        table_data = data.get('data', [])
        
        if not table_data:
            return jsonify({'error': 'No data provided for export'}), 400
        
        # Create Excel file
        excel_file = create_excel_report(table_data)
        
        # Generate filename with timestamp
        filename = f"oracle_table_analysis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Save to backup directory
        backup_path = os.path.join(BACKUP_DIR, filename)
        with open(backup_path, 'wb') as f:
            f.write(excel_file.getvalue())
        
        logger.info(f"Excel file saved to: {backup_path}")
        
        # Reset the BytesIO object for sending
        excel_file.seek(0)
        
        return send_file(
            excel_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        logger.error(f"Error in export_to_excel: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/compare', methods=['POST'])
def compare_files():
    """Compare two Excel files and generate differences report"""
    try:
        # Check if files are uploaded
        if 'file1' not in request.files or 'file2' not in request.files:
            return jsonify({'error': 'Both files are required'}), 400
        
        file1 = request.files['file1']
        file2 = request.files['file2']
        
        if file1.filename == '' or file2.filename == '':
            return jsonify({'error': 'Both files must be selected'}), 400
        
        # Save uploaded files temporarily
        temp_dir = os.path.join(BACKUP_DIR, 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        
        file1_path = os.path.join(temp_dir, f"temp_file1_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        file2_path = os.path.join(temp_dir, f"temp_file2_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        
        file1.save(file1_path)
        file2.save(file2_path)
        
        # Compare files
        differences = compare_excel_files(file1_path, file2_path)
        
        # Clean up temp files
        os.remove(file1_path)
        os.remove(file2_path)
        
        # Check if files are identical
        if not differences:
            return jsonify({
                'success': True,
                'identical': True,
                'message': 'The two files are identical. No differences found.'
            })
        
        # Create comparison report
        comparison_report = create_comparison_report(differences)
        
        # Save to backup directory
        result_filename = f"result_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        result_path = os.path.join(BACKUP_DIR, result_filename)
        
        with open(result_path, 'wb') as f:
            f.write(comparison_report.getvalue())
        
        logger.info(f"Comparison report saved to: {result_path}")
        
        # Reset the BytesIO object for sending
        comparison_report.seek(0)
        
        return send_file(
            comparison_report,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=result_filename
        )
        
    except Exception as e:
        logger.error(f"Error in compare_files: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        connection = get_db_connection()
        connection.close()
        return jsonify({'status': 'healthy', 'database': 'connected'})
    except Exception as e:
        return jsonify({'status': 'unhealthy', 'error': str(e)}), 500
    
    
@app.route('/ui')
def ui():
    """Serve the UI"""
    return render_template('index.html')

@app.route('/analyze')
def analyze_page():
    """Serve the analyze tables page"""
    return render_template('analyze.html')

@app.route('/compare-files')
def compare_files_page():
    """Serve the compare files page"""
    return render_template('compare-files.html')

@app.route('/source-target')
def source_target_page():
    """Serve the source-target comparison page"""
    return render_template('source-target-new.html')

@app.route('/data-comparison')
def data_comparison_page():
    """Serve the SQL query comparison page"""
    return render_template('data-comparison.html')

@app.route('/api/compare-source-target', methods=['POST'])
def compare_source_target():
    """
    Compare source and target tables
    Returns differences in structure, row count, indexes, and constraints
    """
    try:
        data = request.get_json()
        source = data.get('source', {})
        target = data.get('target', {})
        options = data.get('options', {})
        
        # Parse source table with catalog support
        source_input = source.get('table', '').strip()
        source_catalog_or_db, source_schema_parsed, source_table = parse_table_input(source_input)
        
        # Allow explicit schema override
        if source.get('schema') and source.get('schema').strip():
            source_schema = source.get('schema').strip().upper()
        else:
            source_schema = source_schema_parsed
        
        # Parse target table with catalog support
        target_input = target.get('table', '').strip()
        target_catalog_or_db, target_schema_parsed, target_table = parse_table_input(target_input)
        
        # Allow explicit schema override
        if target.get('schema') and target.get('schema').strip():
            target_schema = target.get('schema').strip().upper()
        else:
            target_schema = target_schema_parsed
        
        if not source_table or not target_table:
            return jsonify({'success': False, 'error': 'Source and target table names are required'}), 400
        
        connection = get_db_connection()
        cursor = connection.cursor()
        
        differences = {
            'total_count': 0,
            'structure': [],
            'row_count': {},
            'indexes': [],
            'constraints': []
        }
        
        # Compare structure
        if options.get('structure', True):
            source_structure = get_table_structure(cursor, source_table, source_schema, source_catalog_or_db)
            target_structure = get_table_structure(cursor, target_table, target_schema, target_catalog_or_db)
            
            source_cols = {col['column_name']: col for col in source_structure}
            target_cols = {col['column_name']: col for col in target_structure}
            
            # Find differences
            all_columns = set(list(source_cols.keys()) + list(target_cols.keys()))
            
            for col_name in all_columns:
                if col_name not in source_cols:
                    differences['structure'].append({
                        'column_name': col_name,
                        'diff_type': 'Missing in Source',
                        'source_value': None,
                        'target_value': f"{target_cols[col_name]['data_type']}({target_cols[col_name].get('data_length', '')})",
                        'type': 'added'
                    })
                    differences['total_count'] += 1
                elif col_name not in target_cols:
                    differences['structure'].append({
                        'column_name': col_name,
                        'diff_type': 'Missing in Target',
                        'source_value': f"{source_cols[col_name]['data_type']}({source_cols[col_name].get('data_length', '')})",
                        'target_value': None,
                        'type': 'removed'
                    })
                    differences['total_count'] += 1
                else:
                    # Check if data type or length differs
                    source_col = source_cols[col_name]
                    target_col = target_cols[col_name]
                    
                    if (source_col['data_type'] != target_col['data_type'] or 
                        source_col.get('data_length') != target_col.get('data_length') or
                        source_col.get('data_precision') != target_col.get('data_precision') or
                        source_col.get('nullable') != target_col.get('nullable')):
                        
                        differences['structure'].append({
                            'column_name': col_name,
                            'diff_type': 'Modified',
                            'source_value': f"{source_col['data_type']}({source_col.get('data_length', '')}) {source_col.get('nullable', '')}",
                            'target_value': f"{target_col['data_type']}({target_col.get('data_length', '')}) {target_col.get('nullable', '')}",
                            'type': 'modified'
                        })
                        differences['total_count'] += 1
        
        # Compare row count
        if options.get('rowCount', True):
            source_count = get_table_count(cursor, source_table, source_schema, source_catalog_or_db)
            target_count = get_table_count(cursor, target_table, target_schema, target_catalog_or_db)
            
            source_rows = source_count.get('row_count', 0)
            target_rows = target_count.get('row_count', 0)
            
            differences['row_count'] = {
                'source': source_rows,
                'target': target_rows,
                'difference': abs(source_rows - target_rows) if source_rows and target_rows else 0,
                'different': source_rows != target_rows
            }
            
            if source_rows != target_rows:
                differences['total_count'] += 1
        
        # Compare indexes
        if options.get('indexes', False):
            source_indexes = get_table_indexes(cursor, source_table, source_schema, source_catalog_or_db)
            target_indexes = get_table_indexes(cursor, target_table, target_schema, target_catalog_or_db)
            
            source_idx_names = {idx['index_name'] for idx in source_indexes}
            target_idx_names = {idx['index_name'] for idx in target_indexes}
            
            # Missing in target
            for idx_name in source_idx_names - target_idx_names:
                idx = next((i for i in source_indexes if i['index_name'] == idx_name), None)
                if idx:
                    differences['indexes'].append({
                        'index_name': idx_name,
                        'diff_type': idx.get('index_type', 'INDEX'),
                        'status': 'missing_in_target'
                    })
                    differences['total_count'] += 1
            
            # Missing in source
            for idx_name in target_idx_names - source_idx_names:
                idx = next((i for i in target_indexes if i['index_name'] == idx_name), None)
                if idx:
                    differences['indexes'].append({
                        'index_name': idx_name,
                        'diff_type': idx.get('index_type', 'INDEX'),
                        'status': 'missing_in_source'
                    })
                    differences['total_count'] += 1
        
        # Compare constraints
        if options.get('constraints', False):
            source_pk = get_primary_key(cursor, source_table, source_schema, source_catalog_or_db)
            target_pk = get_primary_key(cursor, target_table, target_schema, target_catalog_or_db)
            source_fks = get_foreign_keys(cursor, source_table, source_schema, source_catalog_or_db)
            target_fks = get_foreign_keys(cursor, target_table, target_schema, target_catalog_or_db)
            
            # Compare primary keys
            has_source_pk = bool(source_pk.get('constraint_name'))
            has_target_pk = bool(target_pk.get('constraint_name'))
            
            if has_source_pk and not has_target_pk:
                # Primary key exists in source but not in target
                differences['constraints'].append({
                    'constraint_name': source_pk['constraint_name'],
                    'constraint_type': 'PRIMARY KEY',
                    'columns': source_pk.get('columns', ''),
                    'source_value': source_pk.get('columns', ''),
                    'target_value': 'None',
                    'status': 'missing_in_target'
                })
                differences['total_count'] += 1
            elif not has_source_pk and has_target_pk:
                # Primary key exists in target but not in source
                differences['constraints'].append({
                    'constraint_name': target_pk['constraint_name'],
                    'constraint_type': 'PRIMARY KEY',
                    'columns': target_pk.get('columns', ''),
                    'source_value': 'None',
                    'target_value': target_pk.get('columns', ''),
                    'status': 'missing_in_source'
                })
                differences['total_count'] += 1
            elif has_source_pk and has_target_pk:
                # Both have primary keys - check if columns are different
                source_pk_cols = source_pk.get('columns', '').strip()
                target_pk_cols = target_pk.get('columns', '').strip()
                
                if source_pk_cols != target_pk_cols:
                    differences['constraints'].append({
                        'constraint_name': f"{source_pk['constraint_name']} / {target_pk['constraint_name']}",
                        'constraint_type': 'PRIMARY KEY',
                        'columns': 'Different columns',
                        'source_value': source_pk_cols,
                        'target_value': target_pk_cols,
                        'status': 'modified'
                    })
                    differences['total_count'] += 1
            
            # Compare foreign keys - create dictionaries for easier comparison
            source_fks_dict = {fk['constraint_name']: fk for fk in source_fks}
            target_fks_dict = {fk['constraint_name']: fk for fk in target_fks}
            
            all_fk_names = set(source_fks_dict.keys()) | set(target_fks_dict.keys())
            
            for fk_name in all_fk_names:
                source_fk = source_fks_dict.get(fk_name)
                target_fk = target_fks_dict.get(fk_name)
                
                if source_fk and not target_fk:
                    # FK exists in source but not in target
                    differences['constraints'].append({
                        'constraint_name': fk_name,
                        'constraint_type': 'FOREIGN KEY',
                        'columns': source_fk.get('columns', ''),
                        'source_value': f"{source_fk.get('columns', '')} -> {source_fk.get('referenced_table', '')}",
                        'target_value': 'None',
                        'status': 'missing_in_target'
                    })
                    differences['total_count'] += 1
                elif not source_fk and target_fk:
                    # FK exists in target but not in source
                    differences['constraints'].append({
                        'constraint_name': fk_name,
                        'constraint_type': 'FOREIGN KEY',
                        'columns': target_fk.get('columns', ''),
                        'source_value': 'None',
                        'target_value': f"{target_fk.get('columns', '')} -> {target_fk.get('referenced_table', '')}",
                        'status': 'missing_in_source'
                    })
                    differences['total_count'] += 1
                elif source_fk and target_fk:
                    # Both have FK with same name - check if details are different
                    source_cols = source_fk.get('columns', '').strip()
                    target_cols = target_fk.get('columns', '').strip()
                    source_ref_table = source_fk.get('referenced_table', '').strip()
                    target_ref_table = target_fk.get('referenced_table', '').strip()
                    
                    if source_cols != target_cols or source_ref_table != target_ref_table:
                        differences['constraints'].append({
                            'constraint_name': fk_name,
                            'constraint_type': 'FOREIGN KEY',
                            'columns': 'Different structure',
                            'source_value': f"{source_cols} -> {source_ref_table}",
                            'target_value': f"{target_cols} -> {target_ref_table}",
                            'status': 'modified'
                        })
                        differences['total_count'] += 1
        
        cursor.close()
        
        # Don't close cached connections (Azure AD/SSO) to prevent re-authentication
        session_id = session.get('session_id')
        if not (session_id and session_id in _connection_cache):
            connection.close()
        
        # Build full table names for display
        db_type = get_db_type()
        source_full_name = build_full_table_name(source_table, source_schema, source_catalog_or_db, db_type)
        target_full_name = build_full_table_name(target_table, target_schema, target_catalog_or_db, db_type)
        
        return jsonify({
            'success': True,
            'source_table': source_full_name,
            'target_table': target_full_name,
            'differences': differences
        })
        
    except Exception as e:
        logger.error(f"Error in compare_source_target: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/compare-query-data', methods=['POST'])
def compare_query_data():
    """Execute and compare SQL query results"""
    try:
        request_data = request.get_json()
        source_sql = request_data.get('source_query', '').strip()
        target_sql = request_data.get('target_query', '').strip()
        
        if not source_sql or not target_sql:
            return jsonify({'success': False, 'error': 'Both queries required'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Execute source query
        cursor.execute(source_sql)
        source_rows = cursor.fetchall()
        source_columns = [desc[0] for desc in cursor.description]
        
        # Execute target query
        cursor.execute(target_sql)
        target_rows = cursor.fetchall()
        target_columns = [desc[0] for desc in cursor.description]
        
        # Compare results (do this BEFORE closing connection in case we need it)
        row_diffs = []
        missing_target = []
        missing_source = []
        matching = 0
        
        # Convert to comparable format
        source_data = [dict(zip(source_columns, row)) for row in source_rows]
        target_data = [dict(zip(target_columns, row)) for row in target_rows]
        
        # Close cursor and connection immediately after fetching data
        cursor.close()
        
        # Don't close cached connections (Azure AD/SSO)
        session_id = session.get('session_id')
        if not (session_id and session_id in _connection_cache):
            conn.close()
        
        # Compare row by row
        max_rows = max(len(source_data), len(target_data))
        for idx in range(max_rows):
            if idx < len(source_data) and idx < len(target_data):
                source_row = source_data[idx]
                target_row = target_data[idx]
                
                # Compare columns
                all_cols = set(list(source_row.keys()) + list(target_row.keys()))
                row_identical = True
                
                for col in all_cols:
                    src_val = source_row.get(col)
                    tgt_val = target_row.get(col)
                    
                    if str(src_val) != str(tgt_val):
                        row_identical = False
                        row_diffs.append({
                            'row_number': idx + 1,
                            'column_name': col,
                            'source_value': str(src_val) if src_val is not None else 'NULL',
                            'target_value': str(tgt_val) if tgt_val is not None else 'NULL'
                        })
                
                if row_identical:
                    matching += 1
            elif idx < len(source_data):
                missing_target.append(idx + 1)
            else:
                missing_source.append(idx + 1)
        
        response = {
            'success': True,
            'summary': {
                'source_rows': len(source_rows),
                'target_rows': len(target_rows),
                'matching_rows': matching,
                'total_differences': len(row_diffs) + len(missing_target) + len(missing_source)
            },
            'differences': {
                'row_differences': row_diffs,
                'missing_in_target': missing_target,
                'missing_in_source': missing_source
            }
        }
        
        return jsonify(response)
        
    except Exception as error:
        logger.error(f"Query comparison error: {str(error)}")
        return jsonify({'success': False, 'error': str(error)}), 500

@app.route('/api/export-query-comparison', methods=['POST'])
def export_query_comparison():
    """Export query comparison to Excel"""
    try:
        comp_data = request.get_json()
        wb = Workbook()
        
        wb.remove(wb.active)
        
        # Summary sheet
        ws_sum = wb.create_sheet("Summary")
        ws_sum.append(["SQL Query Comparison Report"])
        ws_sum.append(["Generated", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        ws_sum.append([])
        ws_sum.append(["Source Rows", comp_data['summary']['source_rows']])
        ws_sum.append(["Target Rows", comp_data['summary']['target_rows']])
        ws_sum.append(["Matching Rows", comp_data['summary']['matching_rows']])
        ws_sum.append(["Total Differences", comp_data['summary']['total_differences']])
        
        # Differences sheet
        if comp_data['differences']['row_differences']:
            ws_diff = wb.create_sheet("Differences")
            ws_diff.append(["Row", "Column", "Source Value", "Target Value"])
            for diff in comp_data['differences']['row_differences']:
                ws_diff.append([
                    diff['row_number'],
                    diff['column_name'],
                    diff['source_value'],
                    diff['target_value']
                ])
        
        output_file = io.BytesIO()
        wb.save(output_file)
        output_file.seek(0)
        
        return send_file(
            output_file,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"query_comparison_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
    except Exception as error:
        logger.error(f"Export error: {str(error)}")
        return jsonify({'success': False, 'error': str(error)}), 500

@app.route('/api/nl-to-sql', methods=['POST'])
def nl_to_sql():
    """
    Natural Language to SQL Generator
    Converts natural language descriptions into SQL queries for multiple databases
    Supports 0-shot, 1-shot, and few-shot learning modes
    """
    try:
        data = request.get_json()
        nl_prompt = data.get('nl_prompt', '').strip()
        learning_mode = data.get('learning_mode', 'zero-shot')
        use_ai = data.get('use_ai', True)
        target_databases = data.get('target_databases', {
            'oracle': True,
            'databricks': True,
            'snowflake': True
        })
        example_prompt = data.get('example_prompt', '').strip()
        example_sql = data.get('example_sql', '').strip()
        
        # Validation
        if not nl_prompt:
            return jsonify({'success': False, 'error': 'No natural language prompt provided'}), 400
        
        if not AI_MODEL_AVAILABLE:
            return jsonify({
                'success': False,
                'error': 'AI model is not configured. Please set AI_API_KEY environment variable.',
                'ai_configured': False
            }), 400
        
        # Check if at least one database is selected
        if not any(target_databases.values()):
            return jsonify({'success': False, 'error': 'Please select at least one target database'}), 400
        
        # Rate limiting check
        rate_limit_check = check_rate_limit()
        if not rate_limit_check['allowed']:
            return jsonify({
                'success': False,
                'error': f"Rate limit exceeded. {rate_limit_check['message']}"
            }), 429
        
        # Generate SQL using AI
        logger.info(f"Generating SQL for prompt: {nl_prompt[:100]}...")
        logger.info(f"Learning mode: {learning_mode}, AI Provider: {AI_MODEL_PROVIDER}")
        
        generated_sql = {}
        ai_explanation = ""
        
        # Generate SQL for each selected database
        for db_type in ['oracle', 'databricks', 'snowflake']:
            if target_databases.get(db_type, False):
                # Build prompt based on learning mode
                system_prompt = f"""You are an expert SQL developer. Convert the natural language description into a valid {db_type.upper()} SQL query.

DATABASE-SPECIFIC SYNTAX:
- Oracle: Use Oracle SQL syntax, DATE functions like TO_DATE(), SYSDATE, ROWNUM for limits
- Databricks: Use Spark SQL syntax, date functions like current_date(), LIMIT for row limits
- Snowflake: Use Snowflake SQL syntax, date functions like CURRENT_DATE(), LIMIT for row limits

IMPORTANT RULES:
1. Generate ONLY the SQL query, no explanations or markdown
2. Use proper {db_type.upper()} syntax
3. Include appropriate WHERE clauses for filtering
4. Use JOINs when multiple tables are involved
5. Add ORDER BY when sorting is mentioned
6. Use aggregate functions (COUNT, SUM, AVG) when appropriate
7. DO NOT use markdown code blocks or backticks
8. Return only raw SQL code"""
                
                user_prompt = f"Convert this to {db_type.upper()} SQL:\n{nl_prompt}"
                
                # Add examples for 1-shot or few-shot learning
                if learning_mode in ['one-shot', 'few-shot'] and example_prompt and example_sql:
                    user_prompt = f"""Here is an example:
Natural Language: {example_prompt}
SQL: {example_sql}

Now convert this to {db_type.upper()} SQL:
{nl_prompt}"""
                
                # Call AI model
                try:
                    if AI_MODEL_PROVIDER == 'openai':
                        # OpenAI API call
                        headers = {
                            'Content-Type': 'application/json',
                            'Authorization': f'Bearer {AI_API_KEY}'
                        }
                        payload = {
                            'model': OPENAI_MODEL,
                            'messages': [
                                {'role': 'system', 'content': system_prompt},
                                {'role': 'user', 'content': user_prompt}
                            ],
                            'temperature': 0.3,
                            'max_tokens': 500
                        }
                        
                        response = requests.post(
                            OPENAI_API_URL,
                            headers=headers,
                            json=payload,
                            timeout=30,
                            verify=False
                        )
                        
                        if response.status_code == 200:
                            result = response.json()
                            sql_query = result['choices'][0]['message']['content'].strip()
                            # Clean up markdown if present
                            sql_query = sql_query.replace('```sql', '').replace('```', '').strip()
                            generated_sql[db_type] = sql_query
                            
                            # Record API call for rate limiting
                            record_api_call()
                        else:
                            logger.error(f"OpenAI API error: {response.status_code} - {response.text}")
                            generated_sql[db_type] = f"-- Error generating {db_type.upper()} SQL"
                    
                    elif AI_MODEL_PROVIDER == 'gemini':
                        # Google Gemini API call
                        api_url = f"{GEMINI_API_URL}?key={AI_API_KEY}"
                        headers = {'Content-Type': 'application/json'}
                        
                        # Combine system and user prompts for Gemini
                        combined_prompt = f"{system_prompt}\n\n{user_prompt}"
                        
                        payload = {
                            'contents': [{
                                'parts': [{'text': combined_prompt}]
                            }],
                            'generationConfig': {
                                'temperature': 0.3,
                                'maxOutputTokens': 500
                            }
                        }
                        
                        response = requests.post(
                            api_url,
                            headers=headers,
                            json=payload,
                            timeout=30,
                            verify=False
                        )
                        
                        if response.status_code == 200:
                            result = response.json()
                            if 'candidates' in result and len(result['candidates']) > 0:
                                sql_query = result['candidates'][0]['content']['parts'][0]['text'].strip()
                                # Clean up markdown if present
                                sql_query = sql_query.replace('```sql', '').replace('```', '').strip()
                                generated_sql[db_type] = sql_query
                                
                                # Record API call for rate limiting
                                record_api_call()
                            else:
                                logger.error(f"Gemini API returned no candidates")
                                generated_sql[db_type] = f"-- Error: No SQL generated for {db_type.upper()}"
                        else:
                            logger.error(f"Gemini API error: {response.status_code} - {response.text}")
                            generated_sql[db_type] = f"-- Error generating {db_type.upper()} SQL: {response.status_code}"
                
                except Exception as e:
                    logger.error(f"Error generating SQL for {db_type}: {str(e)}")
                    generated_sql[db_type] = f"-- Error: {str(e)}"
        
        # Generate AI explanation (only once, not per database)
        if generated_sql and AI_MODEL_PROVIDER == 'gemini':
            try:
                explanation_prompt = f"""Briefly explain what this SQL query does in 1-2 sentences:
{list(generated_sql.values())[0]}"""
                
                api_url = f"{GEMINI_API_URL}?key={AI_API_KEY}"
                payload = {
                    'contents': [{
                        'parts': [{'text': explanation_prompt}]
                    }],
                    'generationConfig': {
                        'temperature': 0.5,
                        'maxOutputTokens': 150
                    }
                }
                
                response = requests.post(api_url, headers={'Content-Type': 'application/json'}, json=payload, timeout=15, verify=False)
                if response.status_code == 200:
                    result = response.json()
                    if 'candidates' in result and len(result['candidates']) > 0:
                        ai_explanation = result['candidates'][0]['content']['parts'][0]['text'].strip()
            except:
                pass  # Explanation is optional
        
        return jsonify({
            'success': True,
            'generated_sql': generated_sql,
            'ai_used': use_ai,
            'ai_provider': AI_MODEL_PROVIDER,
            'learning_mode': learning_mode,
            'explanation': ai_explanation,
            'nl_prompt': nl_prompt
        })
        
    except Exception as e:
        logger.error(f"Error in nl-to-sql: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500

@app.route('/api/export-nl-to-sql', methods=['POST'])
def export_nl_to_sql():
    """Export generated SQL queries to a text file"""
    try:
        data = request.get_json()
        
        # Build export content
        content = "=" * 80 + "\n"
        content += "NATURAL LANGUAGE TO SQL - GENERATED QUERIES\n"
        content += "=" * 80 + "\n\n"
        content += f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        content += f"AI Provider: {data.get('ai_provider', 'Unknown').upper()}\n"
        content += f"Learning Mode: {data.get('learning_mode', 'Unknown').upper()}\n\n"
        
        content += "NATURAL LANGUAGE PROMPT:\n"
        content += "-" * 80 + "\n"
        content += data.get('nl_prompt', '') + "\n\n"
        
        if data.get('explanation'):
            content += "AI EXPLANATION:\n"
            content += "-" * 80 + "\n"
            content += data.get('explanation', '') + "\n\n"
        
        # Add SQL for each database
        generated_sql = data.get('generated_sql', {})
        db_names = {
            'oracle': 'ORACLE SQL',
            'databricks': 'DATABRICKS SQL (Spark SQL)',
            'snowflake': 'SNOWFLAKE SQL'
        }
        
        for db_type in ['oracle', 'databricks', 'snowflake']:
            if db_type in generated_sql:
                content += "=" * 80 + "\n"
                content += f"{db_names[db_type]}\n"
                content += "=" * 80 + "\n"
                content += generated_sql[db_type] + "\n\n"
        
        content += "=" * 80 + "\n"
        content += "END OF GENERATED SQL\n"
        content += "=" * 80 + "\n"
        
        # Return as downloadable file
        return content, 200, {
            'Content-Type': 'text/plain',
            'Content-Disposition': f'attachment; filename=nl_to_sql_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        }
        
    except Exception as e:
        logger.error(f"Export error: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)